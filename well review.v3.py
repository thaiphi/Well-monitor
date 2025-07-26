# well_review.py – Daily Well‐Performance Dashboard (with clickable Well Name, logo, and night mode)
# -----------------------------------------------------------------------------------------
# • Accepts .csv, .xls, .xlsx  (upload or auto‐scan data/)
# • Excel: first sheet, header row=5, report date from AK2
# • For .xlsx: extracts true hyperlinks from “Links” via openpyxl
# • For .xls: reads with pandas/xlrd and sets Link URL = None
# • Aggregates last 3 days, builds flags & Terrible‐Performance score
# • AG‐Grid with pinned “Well Name” & “TerribleScore”
# • Wide layout, resizable columns, color‐coded cells
# • New columns: Running Days, Drive Type, State Detail/Op Mode,
#   Fault Count (cumulative), NearUnderload Ratio & “✗” flag
# • “Normal_vs_Overload” shows a red “✗” (no background)
# • Dashboard <-> Raw‐data page toggle
# • Top‐corner: company logo + contact info
# • Night mode toggle in sidebar

import pathlib, re, datetime as dt, numpy as np, pandas as pd, streamlit as st
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
import openpyxl
import zipfile, xlrd
from io import BytesIO
import json
import urllib.parse
import requests
SETTINGS_FILE = pathlib.Path("customer_settings.json")
# ─── PLACEHOLDER: put your real n8n webhook URL here ───────────────
N8N_WEBHOOK_URL = "https://<YOUR-N8N-HOST>/webhook"
# ───── Load or initialize settings ─────
if SETTINGS_FILE.exists():
    try:
        settings = json.loads(SETTINGS_FILE.read_text())
        if "DEFAULT" not in settings:
            settings["DEFAULT"] = {}
    except Exception as e:
        st.warning(f"⚠️ Failed to read {SETTINGS_FILE.name}. Resetting to blank.")
        settings = {"DEFAULT": {}}
else:
    settings = {"DEFAULT": {}}

# ───── Persisted slider keys (for JSON) ─────
THRESH_KEYS = [
    "CapLoadPct","RiskPct","HighIntake","SmallDrawdown",
    "NearUnderLower","LowUptime","HighDT","VibHigh",
    "ampSpreadRatio","TempHigh","LowDelta","HighFaultCount","HighRunningDays" 
]
WEIGHT_KEYS = [
    "uptime","missing","spread","motortemp","drawdown",
    "delta_tc","nearunderload","vibration","fault"
]

# ───────────────────────── Page config & theme toggle ─────────────────────────
st.set_page_config(layout="wide")

DATA_DIR      = pathlib.Path("data")
LOOKBACK_DAYS = 4   # today + previous 3 days
ROLL_DAYS     = 3
today         = dt.date.today()
csv_date_re   = re.compile(r"(\d{4}-\d{2}-\d{2})", re.I)


# ──────────── Sidebar: Night Mode Toggle & CSS Overrides ─────────────────
night_mode = st.sidebar.checkbox("Night mode")
if night_mode:
    st.markdown(
        """
        <style>
        /* ─── 1) Overall page + block container ───────────────────────────────── */
        /* Make the main body background nearly black, with light text */
        [data-testid="stAppViewContainer"] {
            background-color: #1e1e1e !important;
            color: #e0e0e0 !important;
        }
        /* If you still see .block-container used, override that too: */
        .block-container {
            background-color: #1e1e1e !important;
            color: #e0e0e0 !important;
        }

        /* ─── 2) Sidebar background, headings, and text ───────────────────────── */
        [data-testid="stSidebar"] {
            background-color: #252526 !important;
            color: #e0e0e0 !important;
        }
        /* Force every piece of text inside the sidebar to be light */
        [data-testid="stSidebar"] * {
            color: #e0e0e0 !important;
        }
        /* Sidebar headers (section titles) */
        .css-1d391kg, .css-1v3fvcr, .css-10trblm {
            /* These class names can vary by Streamlit version. 
               If your headings remain dark, inspect and override them here. */
            color: #e0e0e0 !important;
        }

        /* ─── 3) Main headings & markdown text ─────────────────────────────── */
        h1, h2, h3, h4, h5, h6 {
            color: #ffffff !important;
        }
        /* Standard markdown text */
        .css-901oao, .css-16huue1, .css-1n76uvr {
            /* These classes are for normal paragraph text. */
            color: #e0e0e0 !important;
        }
        /* Links inside markdown */
        a {
            color: #66CDAA !important;
        }

        /* ─── 4) Streamlit widgets: buttons, inputs, selects, checkboxes ───────── */
        /* Buttons */
        .stButton>button {
            background-color: #3a3d41 !important;
            color: #ffffff !important;
            border: 1px solid #555 !important;
        }
        .stButton>button:hover {
            background-color: #50545a !important;
        }
        /* Text input boxes and number_input boxes */
        input[type="text"], input[type="number"] {
            background-color: #2d2d2d !important;
            color: #e0e0e0 !important;
            border: 1px solid #555 !important;
        }
        /* Selectboxes, Radio buttons, Checkboxes */
        .stSelectbox, .stRadio, .stCheckbox {
            background-color: #2d2d2d !important;
            color: #e0e0e0 !important;
        }
        /* The actual <label> text for slider, checkbox, etc. */
        label, .css-1v0mbdj, .css-1f4mp12 {
            color: #e0e0e0 !important;
        }
        /* Slider track & handle */
        div[data-baseweb="slider"] {
            background-color: #2d2d2d !important;
        }
        div[data-baseweb="slider"] .rc-slider-rail {
            background-color: #555 !important;
        }
        div[data-baseweb="slider"] .rc-slider-track {
            background-color: #66CDAA !important;
        }
        div[data-baseweb="slider"] .rc-slider-handle {
            background-color: #66CDAA !important;
            border: 1px solid #444 !important;
        }

        /* ─── 5) DataFrame (st.dataframe) override ────────────────────────────── */
        .stDataFrame table {
            background-color: #2d2d2d !important;
            color: #e0e0e0 !important;
        }
        .stDataFrame th {
            background-color: #333 !important;
            color: #e0e0e0 !important;
        }
        .stDataFrame td {
            border-color: #555 !important;
        }

        /* ─── 6) AG-Grid container & cells ───────────────────────────────────── */
        /* (a) Header row: keep dark gray + white text */
        .ag-theme-alpine-dark .ag-header,
        .ag-theme-alpine-dark .ag-header-cell,
        .ag-theme-alpine-dark .ag-header-row {
            background-color: #333 !important;
            color: #e0e0e0 !important;
            border-bottom: 1px solid #555 !important;
        }

        /* (b) Default all data-row cells to white + black text.
            Any inline `js_color` background (green/red) will override this white where it applies. */
        .ag-theme-alpine-dark .ag-center-cols-container .ag-row .ag-cell {
            background-color: #ffffff;    /* plain white */
            color: #000000;               /* plain black */
        }

        /* (c) “Well Name” column cells: force them back to a dark background
            so that your hyperlink‐color logic (blue/red) stands out on dark. */
        .ag-theme-alpine-dark .ag-center-cols-container .ag-cell[col-id="Well Name"] {
            background-color: #2d2d2d !important;
            /* Note: we are NOT forcing a text color here, so your existing
            JS‐renderer (blue link or red “no link”) will still apply. */
        }

        /* (d) Inside “Well Name”, we let <a> tags keep whatever color they already had.
            If you need a custom link‐color, you can add it here. */
        .ag-theme-alpine-dark .ag-center-cols-container .ag-cell[col-id="Well Name"] a {
            /* do NOT override color—your JS renderer’s <a> color stays. */
        }

        /* (e) If your grid has filter/sort icons in the header, brighten them so they remain visible. */
        .ag-theme-alpine-dark .ag-icon {
            filter: brightness(200%) !important;
        }

        /* (f) Ensure the grid’s wrapper behind the header is dark,
            so you don’t see any stray white flashes around the edges. */
        .ag-root-wrapper {
            background-color: #2d2d2d !important;
        }

        /* (g) (Optional) Make the scroll‐thumb a bit darker to blend with night mode */
        .ag-theme-alpine-dark ::-webkit-scrollbar-thumb {
            background-color: #888 !important;
        }
        // ─── Streamlit File Uploader ───────────────────────────── */
        [data-testid="stFileUploader"] {
            background-color: transparent !important;
            color: #ffffff !important;
        }

        [data-testid="stFileUploader"] section {
            background-color: #2d2d2d !important;
            border: 1px solid #555 !important;
            border-radius: 8px !important;
        }

        [data-testid="stFileUploader"] section * {
            color: #ffffff !important;
        }

        [data-testid="stFileDropzone"] {
            background-color: #2d2d2d !important;
            border: 2px dashed #888 !important;
            color: #ffffff !important;
        }

        [data-testid="stFileDropzone"] * {
            color: #ffffff !important;
        }
        [data-testid="stFileUploader"] button {
            background-color: #3a3d41 !important;
            color: #ffffff !important;
            border: 1px solid #777 !important;
            border-radius: 6px !important;
        }

        [data-testid="stFileUploader"] button:hover {
            background-color: #50545a !important;
            color: #ffffff !important;
        }
        /* ─── 8) Scrollbars ───────────────────────────────────────────────────── */
        /* Optional: style scrollbars dark */
        ::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }
        ::-webkit-scrollbar-track {
            background: #1e1e1e;
        }
        ::-webkit-scrollbar-thumb {
            background-color: #555;
            border-radius: 4px;
        }
        ::-webkit-scrollbar-thumb:hover {
            background-color: #666;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

# ───────────────────────── Header: Logo & Contact Info ─────────────────────────
col1, col2 = st.columns([1, 4])
with col1:
    # Make sure "company_logo.png" exists in the same folder, or change the filename here
    st.image("company_logo.png", width=120)
with col2:
    st.markdown(
        """
        <p style="font-size:16px; margin-bottom:4px;">
            <strong>Contact:</strong>
            <a href="mailto:optimization@endurancelift.com" style="color:#0066cc;">
                optimization@endurancelift.com
            </a>
        </p>
        <p style="font-size:16px; margin-top:0;">
            <strong>24 hr hotline:</strong> +1 432 224 0583
        </p>
        """,
        unsafe_allow_html=True,
    )

# ───────────── Helper: ensure Date column ─────────────
def ensure_date_column(df: pd.DataFrame, source_name: str, *, excel_date=None):
    if "Date" not in df.columns:
        if excel_date is not None:
            df.insert(0, "Date", pd.to_datetime(excel_date).date())
        else:
            m = csv_date_re.search(source_name)
            day = m.group(1) if m else today.strftime("%Y-%m-%d")
            df.insert(0, "Date", pd.to_datetime(day).date())
    return df


# ───────────── Excel / CSV loaders ─────────────
def load_excel(buf) -> pd.DataFrame:
    """
    Load .xlsx/.xlsm via pandas + openpyxl hyperlinks,
    or .xls via xlrd (v1.2.0) with hyperlinks.
    Always returns a DataFrame with a 'Date' column and 'Link URL'.
    """
    raw = buf.read()
    fname = buf.name.lower()

    # ─── .xlsx / .xlsm ────────────────────────────────────────────────
    if fname.endswith((".xlsx", ".xlsm")):
        # read data (header row=5 → index=4)
        bio = BytesIO(raw)
        df = pd.read_excel(bio, sheet_name=0, header=4)
        df.columns = [str(c).strip() for c in df.columns]

        # extract report date from AK2
        bio.seek(0)
        dt_val = pd.read_excel(bio, sheet_name=0, header=None, usecols="AK", nrows=2).iloc[1, 0]
        df = ensure_date_column(df, fname, excel_date=dt_val)

        # default Link URL
        df["Link URL"] = None

        # attempt to pull hyperlinks
        try:
            wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            # find "link" column in header row 5
            link_idx = next(
                (i for i, cell in enumerate(ws[5], start=1)
                 if isinstance(cell.value, str) and "link" in cell.value.strip().lower()),
                None
            )
            if link_idx:
                links = []
                for row in ws.iter_rows(min_row=6, max_row=6+len(df)-1, min_col=link_idx, max_col=link_idx):
                    cell = row[0]
                    links.append(cell.hyperlink.target if cell.hyperlink else None)
                df["Link URL"] = links
        except (zipfile.BadZipFile, openpyxl.utils.exceptions.InvalidFileException):
            pass

        return df

    # ─── .xls ─────────────────────────────────────────────────────────
    if fname.endswith(".xls"):
        if xlrd.__version__ != "1.2.0":
            raise RuntimeError(f"xlrd version must be 1.2.0, found {xlrd.__version__}")
        raw_book = xlrd.open_workbook(file_contents=raw, formatting_info=True)
        sheet    = raw_book.sheet_by_index(0)

        # header row=5 → index 4
        cols = [str(v).strip() for v in sheet.row_values(4)]
        data = [sheet.row_values(r) for r in range(5, sheet.nrows)]
        df   = pd.DataFrame(data, columns=cols)

        # extract report date from AK2 (row=2, col AK=index 36)
        dt_val = sheet.cell_value(1, 36)
        df = ensure_date_column(df, fname, excel_date=dt_val)

        df["Link URL"] = None
        # find "link" header index
        link_idx = next((i for i, h in enumerate(cols) if isinstance(h, str) and "link" in h.lower()), None)
        if link_idx is not None and hasattr(sheet, "hyperlink_map"):
            links = []
            for r in range(5, sheet.nrows):
                h = sheet.hyperlink_map.get((r, link_idx))
                links.append(h.url_or_path if h else None)
            df["Link URL"] = links

        return df

    # unsupported
    raise ValueError(f"Unsupported Excel type: {fname}")

def load_csv(buf, name) -> pd.DataFrame:
    """Load CSV, ensure Date column, add Link URL placeholder."""
    # THE FIX: Added header=4 to correctly read the CSV.
    df = pd.read_csv(buf, header=4) 
    df = ensure_date_column(df, name)
    df["Link URL"] = None
    return df

# ───────────── File input & landing-page logic ─────────────
print("📂 Starting file import…", flush=True)

st.sidebar.header("CSV / Excel Source")
upl = st.sidebar.file_uploader(
    "⬆️ Upload 1–4 files (.csv, .xls, .xlsx)",
    type=["csv", "xls", "xlsx"],
    accept_multiple_files=True
)

# collect sources: uploads or fallback to DATA_DIR
if upl:
    sources = upl
else:
    print("   🔄 No upload—falling back to data dir scan", flush=True)
    candidates = [f for f in DATA_DIR.glob("*") if f.suffix.lower() in {".csv", ".xls", ".xlsx"}]
    recent = [f for f in candidates
              if (today - dt.date.fromtimestamp(f.stat().st_mtime)).days < LOOKBACK_DAYS]
    if not recent:
        st.error("❌ No recent files found and none uploaded.")
        st.stop()
    sources = sorted(recent, key=lambda f: f.stat().st_mtime)

dfs = []
for src in sources:
    if upl:
        buf, name = src, src.name
    else:
        buf, name = open(src, "rb"), src.name
    try:
        df = load_csv(buf, name) if name.lower().endswith(".csv") else load_excel(buf)
    finally:
        if not upl:
            buf.close()
    print(f"   📥 Loaded `{name}`, columns = {list(df.columns)}", flush=True)
    dfs.append(df)

if not dfs:
    st.error("❌ No files to process. Upload or add files to data dir.")
    st.stop()

# ───────────── Concatenate all input files and standardize headers ─────────────
df_raw = pd.concat(dfs, ignore_index=True)
df_raw.columns = [str(c).strip() for c in df_raw.columns]
# ─── DROP ANY BLANK‐NAMED COLUMNS ─────────────────────────────────────────────
blank_cols = [c for c in df_raw.columns if c == ""]
if blank_cols:
    print(f"DEBUG: Dropping blank columns: {blank_cols}", flush=True)
    df_raw.drop(columns=blank_cols, inplace=True)

# ───────────── Detect “Customer” column and normalize header ─────────────
cust_col = next((col for col in df_raw.columns if col.strip().lower() == "customer"), None)
if cust_col and cust_col != "Customer":
    df_raw.rename(columns={cust_col: "Customer"}, inplace=True)

# — INSERT THIS —
has_customer = "Customer" in df_raw.columns

if has_customer:
    n_notnull = df_raw["Customer"].notnull().sum()
    unique_vals = df_raw["Customer"].dropna().unique()
    print(f"DEBUG: 'Customer' column present. Non-null count: {n_notnull}")
    print(f"DEBUG: Unique Customer values: {unique_vals}")
    if n_notnull == 0:
        st.warning("The 'Customer' column is present but contains only blanks/NaN.")
else:
    print("DEBUG: 'Customer' column NOT FOUND in columns:", df_raw.columns)
    st.warning("No 'Customer' column found after standardization.")
# ───────────── Manual Customer Name (for single-customer files) ─────────────
if not has_customer:
    cust_input = st.sidebar.text_input("Customer name", "")
    if cust_input:
        df_raw["Customer"] = cust_input
        has_customer = True
# ───────────── Set up well‐to‐customer mapping ─────────────
if has_customer:
    well2cust = df_raw.set_index("Well Name")["Customer"].to_dict()
else:
    well2cust = {}

# ───────────── Determine landing page and reset logic ─────────────
default_page = "Customers" if has_customer else "Dashboard"
valid_pages  = (["Customers"] if has_customer else []) + ["Dashboard", "Raw Data"]

# 1) Reset to default_page if the file list changed
file_list = [getattr(s, "name", s) for s in sources]
if (
    "last_files" not in st.session_state
    or st.session_state.last_files != file_list
):
    st.session_state.last_files  = file_list
    st.session_state.view_page   = default_page



# 3) Initialize view_page if missing or invalid
if (
    "view_page" not in st.session_state
    or st.session_state.view_page not in valid_pages
):
    st.session_state.view_page = default_page

# ───────────── Sidebar page selector (single radio) ─────────────
# --- before you ever show the radio ---
if "view_page" not in st.session_state:
    st.session_state.view_page = default_page

# --- now show the radio, *not* bound to session_state ---
page = st.sidebar.radio(
    "View",
    valid_pages,
    index=valid_pages.index(st.session_state.view_page)
)
# ───────────── Ensure selected_customer exists ─────────────
if "selected_customer" not in st.session_state:
    st.session_state.selected_customer = None
# --- and immediately write back to session_state ---
st.session_state.view_page = page
# ───────────── Determine settings‐key for sliders ─────────────
if page == "Customers":
    current_key = "DEFAULT"
else:
    # on Dashboard, use the selected customer or fall back to DEFAULT
    current_key = st.session_state.selected_customer or "DEFAULT"

# pull ALL saved values (thresholds + weights) for that key
# fall back to whatever DEFAULT has
defs = settings.get(current_key, settings["DEFAULT"])

params = st.query_params
# ─── HANDLE “Trigger” BUTTON → n8n call ─────────────────────────────

if "trigger_well" in params:
    well_to_trigger = params["trigger_well"][0]
    try:
        resp = requests.post(
            N8N_WEBHOOK_URL,
            json={"well_name": well_to_trigger},
            timeout=10
        )
        resp.raise_for_status()
        st.session_state["n8n_response"] = resp.json()
        st.session_state["last_triggered_well"] = well_to_trigger
    except Exception as e:
        st.error(f"❌ Error triggering workflow for {well_to_trigger}: {e}")
    # clear param so we don’t auto-retrigger on rerun
    st.set_query_params()
    st.rerun()
if page == "Customers" and "customer" in params:
    cust = params["customer"][0]
    st.session_state.selected_customer = cust
    st.session_state.view_page = "Dashboard"
    st.set_query_params()      # clear the URL
    st.rerun()
# ───────────── Ensure selected_customer exists ─────────────
if "selected_customer" not in st.session_state:
    st.session_state.selected_customer = None


# ───────────── Load threshold defaults for selected customer ─────────────
current_key = st.session_state.selected_customer or "DEFAULT"
defs = settings.get(current_key, settings["DEFAULT"])

# ───────────── Sidebar thresholds & weights ─────────────
st.sidebar.header("Thresholds")
thr = dict(
    CapLoadPct = st.sidebar.slider(
        "Cap load pct (Max / Normal)", 0.50, 5.0,
        defs.get("CapLoadPct", 1.05), 0.01,
        help=(
            "CapLoad = (Max Drive Amps) ÷ (Normal Running Amps).\n"
            "• We flag At_Max_Capacity whenever CapLoad ≥ this slider’s value."
            "• We also use this same threshold to decide if there’s “spare load” in SpeedUp logic."
        )
    ),
    RiskPct = st.sidebar.slider(
        "Risk pct (Max / Overload)", 0.50, 5.0,
        defs.get("RiskPct", 1.05), 0.01,
        help=(
            "CapRisk = (Max Drive Amps) ÷ (Motor Overload).\n"
            "We flag Overload_Risk whenever CapRisk ≥ this slider’s value."
        )
    ),
    HighIntake = st.sidebar.number_input(
        "High intake ψ (psi)", 0, 10000,
        defs.get("HighIntake", 300),
        help=(
            "High intake threshold is used in SpeedUp logic.\n"
            "We require Avg Intake Pressure > [this value] to consider speeding up."
        )
    ),
    SmallDrawdown = st.sidebar.number_input(
        "Small drawdown ψ (psi)", 0, 5000,
        defs.get("SmallDrawdown", 50),
        help=(
            "Drawdown = (Max Intake Pressure – Min Intake Pressure).\n"
            "We require Drawdown < [this value] to consider speeding up."
        )
    ),
    NearUnderLower = st.sidebar.slider(
        "Near-underload lower bound", 1.0, 5.0,
        defs.get("NearUnderLower", 1.43), 0.01,
        help=(
            "NearUnderload Ratio = (Avg Drive Amps) ÷ (Motor Underload).\n"
            "We flag NearUnderload when ratio < [this slider’s value]."
        )
    ),
    LowUptime = st.sidebar.slider(
        "Low uptime % threshold", 0, 100,
        defs.get("LowUptime", 90),
        help=(
            "If (Uptime %) < [this value], we flag LowUptime.\n"
            "Used in TerribleScore and PoorPerformance."
        )
    ),
    HighDT = st.sidebar.number_input(
        "High downtime hrs (3 days)", 0, 72,
        defs.get("HighDT", 6),
        help=(
            "Downtime (Hr) = hours of downtime over the lookback window.\n"
            "If Downtime > [this value], we flag HighDT (for coloring only)."
        )
    ),
    VibHigh = st.sidebar.number_input(
        "High vibration threshold", 0.0, 10.0,
        defs.get("VibHigh", 1.00), 0.01,
        help=(
            "If (Avg Vib X ≥ [this]) or (Avg Vib Y ≥ [this]), we flag HighVib.\n"
            "Used in TerribleScore and PoorPerformance."
        )
    ),
    FreqSpread = st.sidebar.number_input(
        "High Frequency Spread Ratio threshold (unitless)",
        0.0, 10.0,
        defs.get("FreqSpread", 1.0),
        0.01,
        help="Flag when (Max−Min) ÷ Avg Drive Frequency ≥ this value."
    ),
    ampSpreadRatio = st.sidebar.number_input(
        "High Amp Spread Ratio ≥ (unitless)", 0.0, 10.0,
        defs.get("ampSpreadRatio", 1.0), 0.01,
        help=(
            "ampSpreadRatio = (Max Drive Amps – Min Drive Amps) ÷ (Avg Drive Amps).\n"
            "If Min Drive Amps = 0, we display ‘N/A’ and never flag SpreadFlag.\n"
            "We set SpreadFlag when ampSpreadRatio ≥ [this value]."
        )
    ),
    TempHigh = st.sidebar.number_input(
        "High motor temp °F", 0, 500,
        defs.get("TempHigh", 210),
        help=(
            "If (Max Motor Temp) ≥ [this value], we flag HighMotorTemp.\n"
            "Used in TerribleScore and PoorPerformance."
        )
    ),
    LowDelta = st.sidebar.number_input(
        "Low Tub-Casing Δ ψ", -5000, 5000,
        defs.get("LowDelta", 30),
        help=(
            "Δ = (Avg Tubing Pressure) – (Avg Casing Pressure).\n"
            "If Δ ≤ [this value], we flag LowDeltaTC (for coloring only)."
        )
    ),
    HighFaultCount = st.sidebar.number_input(
        "High fault count (cumulative)", 0, 1000,
        defs.get("HighFaultCount", 1),
        help=(
            "If (Fault Count) ≥ [this value], we flag FaultHigh.\n"
            "Used in TerribleScore and PoorPerformance."
        )
    ),
    HighRunningDays = st.sidebar.number_input(
        "High running days",0, 365 * 5,
        defs.get("HighRunningDays", 90),  # default
        help=("If (Running Days) > this, flag as ‘high running days’")
    ),
    PressureDiff = st.sidebar.number_input(
        "Low Pressure difference threshold (psi)",
        -10000.0, 10000.0,
        defs.get("PressureDiff", 0.0),
        step=0.1,
        help="Flag when (Avg Disch Pressure − Avg Intake Pressure) ≤ this threshold."
    ),
    
    
)
# ─── Default lists for PoorPerformance & SpeedUp ───────────────────
poor_defaults = [
    "LowUptime","HighVib","SpreadFlag","HighMotorTemp","FaultHigh"
]
speed_true_defaults = [
    "Avg Intake Pressure > HighIntake",
    "Drawdown < SmallDrawdown",
    "High running days",
    "At_Max_Capacity",
    "Overload_Risk"
]

# ───────────── PoorPerformance & SpeedUp Settings ─────────────
with st.sidebar.expander("PoorPerformance Settings", expanded=True):
    PoorTrue = st.multiselect(
        "Must be TRUE",
        [
            "LowUptime","HighVib","SpreadFlag","HighMotorTemp","FaultHigh",
            "High running days","At_Max_Capacity","Overload_Risk",
            "High Motor Temp","High Downtime","Max Vibration",
            "Amp Spread Ratio","Tubing-Casing Δ","Fault Count","High Frequency Spread Ratio","Low Pressure Difference",
            "Uptime %","NearUnderload Ratio","NearUnderload",
            "Normal_vs_Overload","MissingSensor"
        ],
        default=defs.get("PoorTrue", ["LowUptime","HighVib","SpreadFlag","HighMotorTemp","FaultHigh"]),
        key="PoorTrue",
        help="Any of these TRUE → contributes to PoorPerformance"
    )
    PoorFalse = st.multiselect(
        "Must be FALSE",
        [
            "LowUptime","HighVib","SpreadFlag","HighMotorTemp","FaultHigh",
            "High running days","At_Max_Capacity","Overload_Risk",
            "High Motor Temp","High Downtime","Max Vibration",
            "Amp Spread Ratio","Tubing-Casing Δ","Fault Count","High Frequency Spread Ratio","Low Pressure Difference",
            "Uptime %","NearUnderload Ratio","NearUnderload",
            "Normal_vs_Overload","MissingSensor"
        ],
        default=defs.get("PoorFalse", []),
        key="PoorFalse",
        help="All of these must be FALSE → to qualify as PoorPerformance"
    )
    thr["PoorTrue"]  = PoorTrue
    thr["PoorFalse"] = PoorFalse

with st.sidebar.expander("SpeedUp Settings", expanded=True):
    SpeedTrue = st.multiselect(
        "Must be TRUE",
        [
            "Avg Intake Pressure > HighIntake","Drawdown < SmallDrawdown",
            "High running days","At_Max_Capacity","Overload_Risk",
            "High Motor Temp","High Downtime","Max Vibration",
            "Amp Spread Ratio","Tubing-Casing Δ","Fault Count","High Frequency Spread Ratio","Low Pressure Difference",
            "HighVib","Uptime %","NearUnderload Ratio","NearUnderload",
            "Normal_vs_Overload","MissingSensor"
        ],
        default=defs.get("SpeedTrue", ["Avg Intake Pressure > HighIntake","Drawdown < SmallDrawdown","High running days","At_Max_Capacity","Overload_Risk"]),
        key="SpeedTrue",
        help="Select flags that must evaluate to True"
    )
    SpeedFalse = st.multiselect(
        "Must be FALSE",
        [
            "Avg Intake Pressure > HighIntake","Drawdown < SmallDrawdown",
            "High running days","At_Max_Capacity","Overload_Risk",
            "High Motor Temp","High Downtime","Max Vibration",
            "Amp Spread Ratio","Tubing-Casing Δ","Fault Count","High Frequency Spread Ratio","Low Pressure Difference",
            "HighVib","Uptime %","NearUnderload Ratio","NearUnderload",
            "Normal_vs_Overload","MissingSensor"
        ],
        default=defs.get("SpeedFalse", []),
        key="SpeedFalse",
        help="Select flags that must evaluate to False"
    )
    thr["SpeedTrue"]  = SpeedTrue
    thr["SpeedFalse"] = SpeedFalse

# ───────────── Inject CSS for coloured pills ─────────────
st.sidebar.markdown(
    """
    <style>
      /* PoorTrue pills */
      div.st-key-PoorTrue span[data-baseweb="tag"] {
        background-color: #002062 !important;
        color:           #ffffff !important;
      }
      /* PoorFalse pills */
      div.st-key-PoorFalse span[data-baseweb="tag"] {
        background-color: #f4bb2a !important;
        color:           #000000 !important;
      }
      /* SpeedTrue pills */
      div.st-key-SpeedTrue span[data-baseweb="tag"] {
        background-color: #002062 !important;
        color:           #ffffff !important;
      }
      /* SpeedFalse pills */
      div.st-key-SpeedFalse span[data-baseweb="tag"] {
        background-color: #f4bb2a !important;
        color:           #000000 !important;
      }
    </style>
    """,
    unsafe_allow_html=True,
)

# ───── Reset per-customer settings ─────
if page == "Dashboard" and st.session_state.selected_customer:
    cust = st.session_state.selected_customer
    if cust in settings and cust != "DEFAULT":
        st.sidebar.info(f"⚙️ Custom settings applied for {cust}.")
        if st.sidebar.button(f"Reset settings to default for {cust}"):
            settings.pop(cust, None)
            SETTINGS_FILE.write_text(json.dumps(settings, indent=2))
            st.sidebar.success(f"Settings for {cust} reset to default.")
            st.rerun()

st.sidebar.markdown("---")
st.sidebar.subheader("Weights – Terrible‐Performance Score")
weights = dict(
    uptime        = st.sidebar.slider(
        "Downtime weight", 0.0, 5.0,
        defs.get("uptime", 1.0), 0.1,
        help=(
            "Adds (LowUptime × [this value]) to TerribleScore.\n"
            "LowUptime = 1 if Uptime% < Low uptime threshold, else 0."
        )
    ),
    missing       = st.sidebar.slider(
        "Missing sensor weight", 0.0, 5.0,
        defs.get("missing", 1.0), 0.1,
        help=(
            "Adds (MissingSensor × [this value]) to TerribleScore.\n"
            "MissingSensor = 1 if Avg Motor Amps=0 or Avg Intake Pressure=0 or flat-line in either; else 0."
        )
    ),
    spread        = st.sidebar.slider(
        "Amp Spread Ratio weight", 0.0, 5.0,
        defs.get("spread", 1.0), 0.1,
        help=(
            "Adds (ampSpreadRatio × [this value]) to TerribleScore.\n"
            "ampSpreadRatio = (Max Drive Amps − Min Drive Amps) ÷ (Avg Drive Amps)."
        )
    ),
    motortemp     = st.sidebar.slider(
        "Motor temp weight", 0.0, 5.0,
        defs.get("motortemp", 1.0), 0.1,
        help=(
            "Adds (HighMotorTemp × [this value]) to TerribleScore.\n"
            "HighMotorTemp = 1 if Max Motor Temp ≥ High motor temp threshold, else 0."
        )
    ),
    drawdown      = st.sidebar.slider(
        "Intake drawdown weight", 0.0, 5.0,
        defs.get("drawdown", 1.0), 0.1,
        help=(
            "Adds ([Drawdown < SmallDrawdown] × [this value]) to TerribleScore.\n"
            "Drawdown = Max Intake Pressure − Min Intake Pressure."
        )
    ),
    delta_tc      = st.sidebar.slider(
        "Tub-Cas Δ weight", 0.0, 5.0,
        defs.get("delta_tc", 1.0), 0.1,
        help=(
            "Removed from TerribleScore (set to 0); kept here only for coloring."
        )
    ),
    nearunderload = st.sidebar.slider(
        "NearUnderload weight", 0.0, 5.0,
        defs.get("nearunderload", 1.0), 0.1,
        help=(
            "Adds (NearUnderload × [this value]) to TerribleScore.\n"
            "NearUnderload = 1 if (Avg Drive Amps ÷ Motor Underload) < Near-underload threshold, else 0."
        )
    ),
    vibration     = st.sidebar.slider(
        "Vibration weight", 0.0, 5.0,
        defs.get("vibration", 1.0), 0.1,
        help=(
            "Adds (HighVib × [this value]) to TerribleScore.\n"
            "HighVib = 1 if (Max Vibration ≥ High vibration threshold)."
        )
    ),
    fault         = st.sidebar.slider(
        "Fault count weight", 0.0, 5.0,
        defs.get("fault", 1.0), 0.1,
        help=(
            "Adds (FaultHigh × [this value]) to TerribleScore.\n"
            "FaultHigh = 1 if Fault Count ≥ High fault count threshold, else 0."
        )
    ),
)

# ───── Save settings buttons ─────
if page == "Customers" and st.sidebar.button("Save default settings"):
    settings["DEFAULT"] = thr
    SETTINGS_FILE.write_text(json.dumps(settings, indent=2))
    # ───── Custom Cards (up to 5) ─────
    st.sidebar.success("Default settings saved.")

if (
    page == "Dashboard"
    and st.session_state.selected_customer
    and st.sidebar.button(f"Save settings for {st.session_state.selected_customer}")
):
    settings[st.session_state.selected_customer] = thr
    SETTINGS_FILE.write_text(json.dumps(settings, indent=2))
    st.sidebar.success(f"Settings saved for {st.session_state.selected_customer}.")
# ───────────── Aggregate last 3 distinct days per well ─────────────
last_dates = sorted(df_raw["Date"].unique())[-ROLL_DAYS:]
df_recent  = df_raw[df_raw["Date"].isin(last_dates)].copy()
hist_days  = len(last_dates)
use_flat   = hist_days >= 3

# ── RIGHT BEFORE the pd.to_numeric loop: Sanitize Uptime (%) ──
if "Uptime (%)" in df_recent.columns:
    col = (
        df_recent["Uptime (%)"]
            .astype(str)
            .str.strip()
            .str.replace('%', '', regex=False)
            .str.lower()
            .replace({'': np.nan, 'n/a': np.nan, 'na': np.nan, 'nan': np.nan, '--': np.nan})
    )
    num = pd.to_numeric(col, errors='coerce')
    # If ≤1.05 assume fraction (0–1); if >1.05 assume percent (0–100)
    df_recent["Uptime (%)"] = np.where(num <= 1.05, num, num / 100.0)

# Convert all other columns (except text fields) to numeric
non_txt = {
    "Well Name", "Field", "Installation Date", "Current Status",
    "Pump Type", "Drive Type", "State Detail/Op Mode", "Links",
    "Latest Fault", "Fault Date", "Link URL"
}

for c in df_recent.columns:
    if c not in non_txt and c != "Date":
        col_data = df_recent[c]
        if not (isinstance(col_data, pd.Series) and col_data.ndim == 1):
            print(f"Column {c} is not a Series or is not 1D, got type {type(col_data)} and ndim {getattr(col_data,'ndim',None)}")
        else:
            df_recent[c] = pd.to_numeric(col_data, errors="coerce")

numeric_cols = df_recent.select_dtypes(include="number").columns.tolist()

text_cols    = [c for c in df_recent.columns if c not in numeric_cols + ["Date"]]

agg_dict = {
    c: ["mean","max","min","std"]
    for c in numeric_cols
    if c!="Well Name"
}
# wrap the "first" in a list so pandas will do SeriesGroupBy.first, not DataFrameGroupBy.first
agg_dict.update({
    c: ["first"]
    for c in text_cols
    if c!="Well Name"
})
df3 = df_recent.groupby("Well Name").agg(agg_dict)
df3.columns = ["_".join(c) if isinstance(c, tuple) else c for c in df3.columns]
df3 = df3.reset_index()

def col(base: str, stat: str) -> str:
    return f"{base}_{stat}"

# ───────────── Latest-day Normal vs Overload ─────────────
latest = (
    df_raw.sort_values("Date")
          .groupby("Well Name", as_index=False)
          .tail(1)
          .set_index("Well Name")
)

# coerce to floats, turning any non-numeric into NaN
df3["Latest_Normal"] = pd.to_numeric(
    latest["Normal Running Amps"]
        .reindex(df3["Well Name"])
        .values,
    errors="coerce"
)
df3["Latest_Overload"] = pd.to_numeric(
    latest["Motor Overload"]
        .reindex(df3["Well Name"])
        .values,
    errors="coerce"
)

# now this comparison will work
df3["Normal_vs_Overload"] = df3["Latest_Normal"] >= df3["Latest_Overload"]

# ───────────── Cap-load & risk (flipped) ─────────────
df3["CapLoad"] = df3[col("Max Drive Amps", "mean")] / df3[col("Normal Running Amps", "mean")]
df3["CapRisk"] = df3[col("Max Drive Amps", "mean")] / df3[col("Motor Overload", "mean")]

df3["At_Max_Capacity"] = df3["CapLoad"] >= thr["CapLoadPct"]
df3["Overload_Risk"]   = df3["CapRisk"] >= thr["RiskPct"]

# ───────────── Additional flags and derived columns ─────────────
drawdown = df3[col("Max Intake Pressure", "max")] - df3[col("Min Intake Pressure", "min")]
df3["HighRunningDays"] = df3[col("Running Days", "mean")] > thr["HighRunningDays"]
df3["HighDowntime"]    = df3[col("Downtime (Hr)", "mean")] > thr["HighDT"]
df3["LowDeltaFlag"]    = (
    df3[col("Avg Tubing", "mean")] - df3[col("Avg Casing", "mean")]
) <= thr["LowDelta"]
df3["NearUnderload Ratio"] = df3[col("Avg Drive Amps", "mean")] / df3[col("Motor Underload", "mean")]
df3["NearUnderload"] = df3["NearUnderload Ratio"] < thr["NearUnderLower"]

df3["Max Vibration"] = df3[[col("Avg Vib X", "mean"), col("Avg Vib Y", "mean")]].max(axis=1)
df3["HighVib"] = df3["Max Vibration"] >= thr["VibHigh"]

df3["HighMotorTemp"] = df3[col("Max Motor Temp", "max")] >= thr["TempHigh"]
df3["Pressure Difference"] = (
    df3[col("Avg Disch Pressure", "mean")]
    - df3[col("Avg Intake Pressure", "mean")]
)

freq_range = (
    df3[col("Max Drive Frequency", "max")]
    - df3[col("Min Drive Frequency", "min")]
)
df3["Frequency Spread Ratio"] = np.where(
    df3[col("Avg Drive Frequency", "mean")] > 0,
    freq_range / df3[col("Avg Drive Frequency", "mean")],
    np.nan
)
# ───────────── Compute Amp Spread Ratio & Flag ─────────────
spread_ratio = (
    (df3[col("Max Drive Amps", "max")] - df3[col("Min Drive Amps", "min")])
    / df3[col("Avg Drive Amps", "mean")]
)
df3["ampSpreadRatio"] = spread_ratio.where(df3[col("Min Drive Amps", "min")] != 0, np.nan)
df3["SpreadFlag"] = (
    (df3["ampSpreadRatio"] >= thr["ampSpreadRatio"]) &
    (df3[col("Min Drive Amps", "min")] != 0)
)
df3["Lost_Motor"] = (
    (df3[col("Avg Motor Amps", "mean")] == 0) |
    (use_flat & (df3[col("Avg Motor Amps", "std")] == 0))
)
df3["Lost_Intake"] = (
    (df3[col("Avg Intake Pressure", "mean")] == 0) |
    (use_flat & (df3[col("Avg Intake Pressure", "std")] == 0))
)
df3["MissingSensor"] = df3[["Lost_Motor", "Lost_Intake"]].any(axis=1)

# Uptime %: scale fraction (0–1) → 0–100
uptime_pct       = df3[col("Uptime (%)", "mean")] * 100
df3["LowUptime"] = uptime_pct < thr["LowUptime"]




# ───────────── Fault Count & Flag ─────────────
# Choose the right raw fault-count column (daily vs weekly)
fault_24 = col("Fault Count (24hr)", "mean")
fault_7d = col("Fault Count\n(7 Day)",    "mean")
if fault_24 in df3.columns:
    fault_col = fault_24
elif fault_7d in df3.columns:
    fault_col = fault_7d
else:
    st.warning("No fault-count column found; defaulting to zero.")
    fault_col = None

if fault_col:
    fault_mean = (
        df3[fault_col]
        .replace([np.inf, -np.inf], 0)
        .fillna(0)
    )
else:
    fault_mean = pd.Series(0, index=df3.index)

# then downstream:
df3["Fault Count"] = (fault_mean * hist_days).round().astype(int)
df3["FaultHigh"]   = df3["Fault Count"] >= thr["HighFaultCount"]
df3["ModemOffline"] = (
    df3[col("State Detail/Op Mode", "first")] == "MODEM OFFLINE"
)
# ─── build a map from your sidebar strings → boolean series ─────────
flag_map = {
    "LowUptime":          df3["LowUptime"],
    "HighVib":            df3["HighVib"],
    "SpreadFlag":         df3["SpreadFlag"],
    "HighMotorTemp":      df3["HighMotorTemp"],
    "FaultHigh":          df3["FaultHigh"],

    "High running days":  df3["HighRunningDays"],
    "High Downtime":      df3["HighDowntime"],
    "LowDeltaFlag":       df3["LowDeltaFlag"],

    "At_Max_Capacity":    df3["At_Max_Capacity"],
    "Overload_Risk":      df3["Overload_Risk"],
    "High Motor Temp":    df3["HighMotorTemp"],
    "Max Vibration":      df3["HighVib"],
    "Amp Spread Ratio":       df3["SpreadFlag"],
    "Tubing-Casing Δ":    df3["LowDeltaFlag"],
    "Fault Count":        df3["FaultHigh"],
    "Uptime %":           df3["LowUptime"],
    "NearUnderload Ratio":df3["NearUnderload"],
    "NearUnderload":      df3["NearUnderload"],
    "Normal_vs_Overload": df3["Normal_vs_Overload"],
    "MissingSensor":      df3["MissingSensor"],

    "Avg Intake Pressure > HighIntake":
        df3[col("Avg Intake Pressure","mean")] > thr["HighIntake"],
    "Drawdown < SmallDrawdown":
        drawdown < thr["SmallDrawdown"],
}

# ─── PoorPerformance now = OR over your selected criteria ────────────
# ─── compute PoorPerformance via TRUE & FALSE lists ───────────────────
# build the TRUE‐flags series
pt = thr["PoorTrue"]
if pt:
    true_df = pd.concat([flag_map[c] for c in pt if c in flag_map], axis=1)
    poor_true = true_df.any(axis=1)
else:
    poor_true = pd.Series(False, index=df3.index)

# build the FALSE‐flags series
pf = thr["PoorFalse"]
if pf:
    false_df = pd.concat([flag_map[c] for c in pf if c in flag_map], axis=1)
    poor_false_ok = (~false_df).all(axis=1)
else:
    poor_false_ok = pd.Series(True, index=df3.index)

# final PoorPerformance: any TRUE *and* all FALSE
df3["PoorPerformance"] = poor_true & poor_false_ok



# ─── SpeedUp = AND across all TRUE flags and all FALSE flags ─────────
true_list  = thr["SpeedTrue"]
false_list = thr["SpeedFalse"]
series = []

for c in true_list:
    if c in flag_map:
        series.append(flag_map[c])
for c in false_list:
    if c in flag_map:
        series.append(~flag_map[c])

# if no criteria selected, default to False
if series:
    df3["SpeedUp"] = pd.concat(series, axis=1).all(axis=1)
else:
    df3["SpeedUp"] = False



# ───────────── Terrible‐Performance Score ─────────────
df3["TerribleScore"] = (
      weights["uptime"]       * df3["LowUptime"].astype(int)
    + weights["missing"]      * df3["MissingSensor"].astype(int)
    + weights["spread"]       * df3["ampSpreadRatio"].fillna(0)
    + weights["motortemp"]    * df3["HighMotorTemp"].astype(int)
    + weights["drawdown"]     * (drawdown < thr["SmallDrawdown"]).astype(int)
    + weights["nearunderload"]* df3["NearUnderload"].astype(int)
    + weights["vibration"]    * df3["HighVib"].astype(int)
    + weights["fault"]        * df3["FaultHigh"].astype(int)
)

# ───────────── Build AG‐Grid table ───────────────────────────────────────
df_show = df3.copy()
# ─── INSERT a dummy Trigger column for the grid button ─────────────
df_show["Trigger"] = "Trigger"

# ─── Add hidden Boolean columns for PoorPerformance reasons ───
df_show["LowUptime_bool"]    = df3["LowUptime"]       # True = failed uptime threshold
df_show["HighVib_bool"]      = df3["HighVib"]         # True = failed vibration threshold
df_show["SpreadFlag_bool"]   = df3["SpreadFlag"]      # True = failed spread‐ratio threshold
df_show["HighMotorTemp_bool"]= df3["HighMotorTemp"]   # True = failed motor‐temp threshold
df_show["FaultHigh_bool"]    = df3["FaultHigh"]       # True = failed fault‐count threshold
df_show["HighRunningDays_bool"] = df3["HighRunningDays"]
df_show["HighDowntime_bool"]    = df3["HighDowntime"]
df_show["LowDeltaFlag_bool"]    = df3["LowDeltaFlag"]
# ─── Add hidden Boolean columns for SpeedUp reasons ───
#  1) Running Days < 90?
df_show["Speed_RunDays_OK"]     = df3[col("Running Days", "mean")] < 90

#  2) Avg Intake Pressure > HighIntake?
df_show["Speed_AvgIntake_OK"]   = df3[col("Avg Intake Pressure", "mean")] > thr["HighIntake"]

#  3) Drawdown < SmallDrawdown?
#     (we already computed `drawdown = Max Intake – Min Intake` above)
df_show["Speed_Drawdown_OK"]    = drawdown < thr["SmallDrawdown"]

#  4) Overload_Risk == False?
df_show["Speed_OverloadOK"]     = ~df3["Overload_Risk"].astype(bool)

#  5) At_Max_Capacity == False?
df_show["Speed_AtMaxOK"]        = ~df3["At_Max_Capacity"].astype(bool)

# Numeric columns first
df_show["High Motor Temp"]     = df3[col("Max Motor Temp", "max")]
df_show["High Downtime"]       = df3[col("Downtime (Hr)", "mean")]
df_show["Max Vibration"]       = df3["Max Vibration"]

# Amp Spread Ratio / Tubing-Casing Δ / NearUnderload Ratio
df_show["Amp Spread Ratio"]        = df3["ampSpreadRatio"]
df_show["Tubing-Casing Δ"]     = df3[col("Avg Tubing", "mean")] - df3[col("Avg Casing", "mean")]
df_show["NearUnderload Ratio"] = df3["NearUnderload Ratio"]
df_show["Pressure Difference"] = df3["Pressure Difference"]
df_show["Frequency Spread Ratio"]         = df3["Frequency Spread Ratio"]
# Boolean‐derived columns replaced with numeric or checkmarks:
df_show["Normal_vs_Overload"]  = df3["Normal_vs_Overload"].apply(lambda x: "✗" if x else "")
df_show["MissingSensor"]       = df3["MissingSensor"].apply(lambda x: "✗" if x else "")

df_show["Fault Count"] = df3["Fault Count"]       # numeric, colored below
df_show["HighVib"]     = df3["Max Vibration"]     # numeric, colored below

df_show["Uptime %"]    = uptime_pct               # numeric, colored below
df_show["PoorPerformance"] = df3["PoorPerformance"].apply(lambda x: "✗" if x else "✓")
df_show["SpeedUp"] = df3["SpeedUp"].apply(lambda x: "✓" if x else "✗")

df_show["At_Max_Capacity"] = df3["CapLoad"]
df_show["Overload_Risk"]   = df3["CapRisk"]

df_show["Running Days"]        = df3[col("Running Days", "mean")]
df_show["Drive Type"]          = df3[col("Drive Type", "first")]
df_show["State Detail/Op Mode"]= df3[col("State Detail/Op Mode", "first")]

# Link URL (first value for each well)
df3["Link URL"] = df3[col("Link URL", "first")]
df_show["Link URL"] = df3["Link URL"]

# Columns to display, in order:
display_cols = [
    "Well Name", "Trigger","Running Days", "TerribleScore", "PoorPerformance", "SpeedUp",
    "At_Max_Capacity", "Overload_Risk",
    "High Motor Temp", "High Downtime", "Max Vibration", "Pressure Difference", "Frequency Spread Ratio",
    "Amp Spread Ratio", "Tubing-Casing Δ","Fault Count",
    "Uptime %", "NearUnderload Ratio",
    "Normal_vs_Overload",
    "MissingSensor", "Drive Type", "State Detail/Op Mode"
]
# ───── Fields available for custom cards ─────
# Only include the numeric display columns from df_show
card_fields = df3.select_dtypes(include=[np.number]).columns.tolist()
with st.sidebar.expander("Custom Cards", expanded=False):
    
    custom = settings.get(current_key, {}).get("custom_cards", [])
    # List & remove
    for i, card in enumerate(custom):
        c1, c2 = st.columns([4,1])
        with c1:
            st.markdown(f"**{card['label']}**: {card['field']} {card['operator']} {card['threshold']}")
        with c2:
            if st.button("❌", key=f"rm_card_{i}"):
                settings[current_key]["custom_cards"].pop(i)
                SETTINGS_FILE.write_text(json.dumps(settings, indent=2))
                st.rerun()
    # Add new
    if len(custom) < 5:
        lbl    = st.text_input("Label")
        st.markdown("**Condition 1**")
        field1 = st.selectbox("Field",    card_fields, key="c1f")
        op1     = st.selectbox("Operator", [">","<","="],    key="c1o")
        val1    = st.number_input("Threshold", value=0.0,    key="c1v")

        add2    = st.checkbox("Add second condition?",      key="add2")
        if add2:
            st.markdown("**Condition 2**")
            field2 = st.selectbox("Field (2)", card_fields,  key="c2f")
            op2     = st.selectbox("Operator (2)", [">","<","="], key="c2o")
            val2    = st.number_input("Threshold (2)", value=0.0, key="c2v")
            comb    = st.radio("Combine with", ["AND","OR"], index=0, key="c2c")
        else:
            field2 = op2 = val2 = comb = None

        color  = st.text_input("Color (hex)", "#336699", key="c_color")
        if st.button("Add card"):
            new = {
                "label": lbl,
                "conditions": [
                    {"field": field1, "op": op1, "value": val1}
                ],
                "combiner": comb or "AND",
                "color": color
            }
            if add2:
                new["conditions"].append(
                    {"field": field2, "op": op2, "value": val2}
                )
            custom.append(new)
            SETTINGS_FILE.write_text(json.dumps(settings, indent=2))
            st.rerun()


# ---------------------- JS Cell‐Style + Link‐Renderer ----------------------
js_color = JsCode(f"""
function(p) {{
    // Convert Python night_mode (True/False) → JS boolean (true/false)
    var isDark = {str(night_mode).lower()};  

    // Define color pairs for light vs dark mode
    var lightRed  = '#FFB3B3';
    var lightGreen = '#C6F7C6';
    var darkRed   = '#E74C3C';    // darker red for contrast
    var darkGreen = '#2ECC71';    // emerald green in dark mode
    var cellText  = isDark ? '#FFFFFF' : '#000000';

    // At_Max_Capacity coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'At_Max_Capacity') {{
        var bg = (p.value >= {thr["CapLoadPct"]}) 
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Overload_Risk coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'Overload_Risk') {{
        var bg = (p.value >= {thr["RiskPct"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // High Motor Temp coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'High Motor Temp') {{
        var bg = (p.value !== null && p.value >= {thr["TempHigh"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // High Downtime coloring: red > threshold, green otherwise
    if (p.colDef.field === 'High Downtime') {{
        var bg = (p.value > {thr["HighDT"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Max Vibration coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'Max Vibration') {{
        var bg = (p.value >= {thr["VibHigh"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Amp Spread Ratio coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'Amp Spread Ratio') {{
        var bg = ((p.value !== null) && (p.value >= {thr["ampSpreadRatio"]}))
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}
    if (p.colDef.field === 'Pressure Difference') {{
        var bg = (p.value <= {thr["PressureDiff"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Frequency Spread Ratio coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'Frequency Spread Ratio') {{
        var bg = (p.value >= {thr["FreqSpread"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Tubing-Casing Δ coloring: red ≤ threshold, green otherwise
    if (p.colDef.field === 'Tubing-Casing Δ') {{
        var bg = (p.value <= {thr["LowDelta"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // NearUnderload Ratio coloring: red < threshold, green otherwise
    if (p.colDef.field === 'NearUnderload Ratio') {{
        var bg = (p.value < {thr["NearUnderLower"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // NearUnderload boolean: red “✗” if True
    if (p.colDef.field === 'NearUnderload') {{
        if (p.value === '✗') {{
            return {{ 'color': (isDark ? '#FF6961' : '#FF0000') }};
        }}
        return null;
    }}

    // Normal_vs_Overload: red “✗” if True
    if (p.colDef.field === 'Normal_vs_Overload') {{
        if (p.value === '✗') {{
            return {{ 'color': (isDark ? '#FF6961' : '#FF0000') }};
        }}
        return null;
    }}

    // MissingSensor boolean: red “✗” if True
    if (p.colDef.field === 'MissingSensor') {{
        if (p.value === '✗') {{
            return {{ 'color': (isDark ? '#FF6961' : '#FF0000') }};
        }}
        return null;
    }}

    // Fault Count: red if ≥ threshold, green otherwise
    if (p.colDef.field === 'Fault Count') {{
        var bg = (p.value >= {thr["HighFaultCount"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // HighVib: red if ≥ threshold, green otherwise
    if (p.colDef.field === 'HighVib') {{
        var bg = (p.value >= {thr["VibHigh"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Uptime %: red if < LowUptime threshold, green otherwise
    if (p.colDef.field === 'Uptime %') {{
        var bg = (p.value < {thr["LowUptime"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // PoorPerformance boolean: red “✗” if True (no background)
    if (p.colDef.field === 'PoorPerformance') {{
        if (p.value === '✗') {{
            return {{ 'color': (isDark ? '#FF6961' : '#FF0000') }};
        }}
        return null;
    }}

    // SpeedUp: green “✓” if True (no background)
    if (p.colDef.field === 'SpeedUp') {{
        if (p.value === '✓') {{
            return {{ 'color': (isDark ? '#77DD77' : '#00AA00') }};
        }}
        return null;
    }}

    // If no conditions matched, return null so AG-Grid uses default
    return null;
}}
""")

gb = GridOptionsBuilder.from_dataframe(df_show[display_cols])

# Configure “Well Name” as a clickable hyperlink or red if no URL
gb.configure_column(
    "Well Name",
    pinned="left",
    cellRenderer=JsCode("""
        function (params) {
            const url   = params.data["Link URL"];
            const value = params.value || "";
            if (url) {
                // If there’s a URL, wrap the well name in an <a> tag:
                params.eGridCell.innerHTML =
                    `<a href="${url}" target="_blank">${value}</a>`;
            } else {
                // Otherwise, show the well name in red
                params.eGridCell.innerHTML =
                    `<span style="color:red">${value}</span>`;
            }
            return null;
        }
    """)
)
# ─── Configure our new Trigger button column ──────────────────────────
gb.configure_column(
    "Trigger",
    headerName="",
    pinned="left",
    width=100,
    cellRenderer=JsCode("""
        function(params) {
            const well = params.data["Well Name"];
            const href = window.location.pathname
                       + "?trigger_well="
                       + encodeURIComponent(well);
            // inject a real <button> into the cell
            params.eGridCell.innerHTML = 
                `<button style="width:80px;" 
                         onclick="window.location.href='${href}'">
                     Trigger
                 </button>`;
            return null;
        }
    """)
)

# Pin “Running Days” so it sits to the left of “Well Name”
gb.configure_column(
    "Running Days",
    pinned="left",
    type=["numericColumn"],
    valueFormatter="x.toFixed(1)",
    headerTooltip="Mean Running Days over lookback window"
)

# Pin TerribleScore with tooltip
gb.configure_column(
    "TerribleScore",
    pinned="left",
    type=["numericColumn"],
    headerTooltip=(
        "TerribleScore = "
        "(LowUptime × uptime weight) + "
        "(MissingSensor × missing weight) + "
        "(ampSpreadRatio × spread weight) + "
        "(HighMotorTemp × motortemp weight) + "
        "((Max Intake – Min Intake) < SmallDrawdown threshold × drawdown weight) + "
        "(NearUnderload × nearunderload weight) + "
        "(HighVib × vibration weight) + "
        "(FaultHigh × fault count weight)."
    )
)

poor_true = defs.get("PoorTrue", poor_defaults)
poor_false = defs.get("PoorFalse", [])

gb.configure_column(
    "PoorPerformance",
    pinned="left",
    type=["textColumn"],
    headerTooltip=(
        "Displays ✗ if any TRUE‐flag is met or any FALSE‐flag is violated."
    ),
    tooltipValueGetter=JsCode(f"""
        function(params) {{
            if (params.value === '✓') return null;
            const trueList  = {json.dumps(poor_true)};
            const falseList = {json.dumps(poor_false)};
            const d = params.data;
            let lines = [];
            // TRUE‐flags first
            trueList.forEach(flag => {{
                const ok = d[flag + '_bool'];
                lines.push(`${{flag}} is ${{ok}}: ${{ok ? 'passed' : 'failed'}}`);
            }});
            // then FALSE‐flags
            falseList.forEach(flag => {{
                const ok = d[flag + '_bool'];
                lines.push(`${{flag}} is ${{ok}}: ${{ok ? 'passed' : 'failed'}}`);
            }});
            return lines.join("\\n");
        }}
    """)
)

# pull your saved lists out of defs
speed_true  = defs.get("SpeedTrue", speed_true_defaults)
speed_false = defs.get("SpeedFalse", [])

# pre-dump to JSON
true_list_json  = json.dumps(speed_true)
false_list_json = json.dumps(speed_false)

gb.configure_column(
    "SpeedUp",
    pinned="left",
    type=["textColumn"],
    headerTooltip="Displays ✓ only if all TRUE-flags are met and FALSE-flags are clear.",
    tooltipValueGetter=JsCode(
        """
        function(params) {
            if (params.value === '✓') return null;
            const trueList  = %(true)s;
            const falseList = %(false)s;
            const d = params.data;

            // map your sidebar labels → the actual boolean-column names in `d`
            const map = {
              "Avg Intake Pressure > HighIntake": "Speed_AvgIntake_OK",
              "Drawdown < SmallDrawdown":         "Speed_Drawdown_OK",
              "High running days":                "HighRunningDays_bool",
              "At_Max_Capacity":                  "Speed_AtMaxOK",
              "Overload_Risk":                    "Speed_OverloadOK"
              // add more mappings here if you expose them…
            };

            function describe(flag) {
              const field = map[flag] || (flag + "_bool");
              const val   = d[field];
              const ok    = (typeof val === "boolean") ? val : undefined;
              const status = ok === undefined ? "undefined" : ok;
              const result = ok ? "passed" : "failed";
              return `${flag} is ${status}: ${result}`;
            }

            let lines = [];
            trueList.forEach(f => lines.push(describe(f)));
            falseList.forEach(f => lines.push(describe(f)));
            return lines.join("\\n");
        }
        """ % {
            "true":  true_list_json,
            "false": false_list_json
        }
    )
)
# Configure each derived column’s tooltip & formatting:
gb.configure_column(
    "At_Max_Capacity",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "At_Max_Capacity = Max Drive Amps ÷ Normal Running Amps.\n"
        "Cell is red if ≥ Cap load pct threshold, green otherwise."
    )
)
gb.configure_column(
    "Overload_Risk",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "Overload_Risk = Max Drive Amps ÷ Motor Overload.\n"
        "Cell is red if ≥ Risk pct threshold, green otherwise."
    )
)
gb.configure_column(
    "High Motor Temp",
    type=["numericColumn"],
    valueFormatter="x.toFixed(1)",
    headerTooltip=(
        "High Motor Temp = Max Motor Temp.\n"
        "Cell is red if ≥ High motor temp threshold, green otherwise."
    )
)
gb.configure_column(
    "High Downtime",
    type=["numericColumn"],
    valueFormatter="x.toFixed(1)",
    headerTooltip=(
        "High Downtime = average Downtime (Hr) over lookback window.\n"
        "Cell is red if > High downtime hrs threshold, green otherwise."
    )
)
gb.configure_column(
    "Max Vibration",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "Max Vibration = max(Avg Vib X, Avg Vib Y) over lookback.\n"
        "Cell is red if ≥ High vibration threshold, green otherwise."
    )
)
gb.configure_column(
    "Amp Spread Ratio",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "ampSpreadRatio = (Max Drive Amps − Min Drive Amps) ÷ (Avg Drive Amps).\n"
        "If Min Drive Amps = 0 → displayed as N/A.\n"
        "Cell is red if ≥ Amp Spread Ratio threshold, green otherwise."
    )
)

gb.configure_column(
    "Tubing-Casing Δ",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "Tubing-Casing Δ = Avg Tubing Pressure − Avg Casing Pressure.\n"
        "Cell is red if ≤ Low Tub-Casing Δ threshold, green otherwise."
    )
)
gb.configure_column(
    "Pressure Difference",
    type=["numericColumn"],
    valueFormatter="x == null ? '' : x.toFixed(1)",
    headerTooltip=(
        "Pressure Difference = Avg Disch Pressure − Avg Intake Pressure."
        "Cell is red if ≤ Pressure Difference threshold, green otherwise."
    )
)
# ---------- NEW: Frequency Spread Ratio --------------------------------
gb.configure_column(
    "Frequency Spread Ratio",
    type=["numericColumn"],
    valueFormatter="x == null ? '' : x.toFixed(2)",
    headerTooltip=(
        "Frequency Spread Ratio = (Max − Min) ÷ Avg Drive Frequency (unitless).\n"
        "Cell is red if ≥ Frequency Spread Ratio threshold, green otherwise."
    )
)
gb.configure_column(
    "NearUnderload Ratio",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "NearUnderload Ratio = Avg Drive Amps ÷ Motor Underload.\n"
        "Cell is red if < Near-underload lower bound, green otherwise."
    )
)

gb.configure_column(
    "Normal_vs_Overload",
    type=["textColumn"],
    headerTooltip=(
        "True if (Latest Normal Running Amps) ≥ (Latest Motor Overload), else False.\n"
        "Displays a red “✗” if True, blank if False."
    )
)
gb.configure_column(
    "MissingSensor",
    type=["textColumn"],
    headerTooltip=(
        "True if Avg Motor Amps=0 or Avg Intake Pressure=0 or flat-line in either.\n"
        "Displays a red “✗” if True, blank if False."
    )
)
gb.configure_column(
    "Fault Count",
    type=["numericColumn"],
    valueFormatter="x.toFixed(0)",
    headerTooltip=(
        "Fault Count = cumulative faults over lookback window.\n"
        "Cell is red if ≥ High fault count threshold, green otherwise."
    )
)

gb.configure_column(
    "Uptime %",
    type=["numericColumn"],
    valueFormatter="x == null ? '' : x.toFixed(1)",
    headerTooltip=(
        "Uptime % = mean uptime percentage over lookback window.\n"
        "Cell is red if < Low uptime threshold, green otherwise."
    )
)

gb.configure_column(
    "Drive Type",
    type=["textColumn"],
    headerTooltip="Drive Type (first value over lookback window)."
)
gb.configure_column(
    "State Detail/Op Mode",
    type=["textColumn"],
    headerTooltip="State Detail/Op Mode (first value over lookback window)."
)
# ─── Hide the helper Boolean fields from the visible grid ───
gb.configure_column("LowUptime_bool",       hide=True)
gb.configure_column("HighVib_bool",         hide=True)
gb.configure_column("SpreadFlag_bool",      hide=True)
gb.configure_column("HighMotorTemp_bool",   hide=True)
gb.configure_column("FaultHigh_bool",       hide=True)

gb.configure_column("Speed_RunDays_OK",     hide=True)
gb.configure_column("Speed_AvgIntake_OK",   hide=True)
gb.configure_column("Speed_Drawdown_OK",    hide=True)
gb.configure_column("Speed_OverloadOK",     hide=True)
gb.configure_column("Speed_AtMaxOK",        hide=True)
gb.configure_column("HighRunningDays_bool", hide=True)
gb.configure_column("HighDowntime_bool",    hide=True)
gb.configure_column("LowDeltaFlag_bool",    hide=True)

gb.configure_grid_options(enableBrowserTooltips=True)
gb.configure_default_column(resizable=True, minWidth=120)
gb.configure_grid_options(domLayout='normal')
gb.configure_columns(display_cols, cellStyle=js_color)
grid_opts = gb.build()

# ───────────── Page toggle ─────────────

# ───────────── CUSTOMERS LANDING PAGE ─────────────
if page == "Customers":
    st.title("Customer Overview")

    # map each well → its customer
    df3["Customer"] = df3["Well Name"].map(well2cust)
    df_show["Customer"] = df3["Customer"]
    # ───────────── Precompute drive‐type counts per customer ─────────────
    drive_counts_df = (
        df_show
         .groupby(["Customer", "Drive Type"])
         .size()
         .reset_index(name="Count")
    )
    # Total across all customers
    overall_drive_counts = (
        df_show
        .groupby("Drive Type")
        .size()
        .reset_index(name="Count")
    )
    with st.expander("Drive types chart", expanded=False):
        drive_df = overall_drive_counts.sort_values("Count", ascending=False)
        st.bar_chart(drive_df.set_index("Drive Type")["Count"])




    # build your metrics, including per-customer well count
    cust_metrics = (
        df3.groupby("Customer")
           .agg(
               WellCount             = ("Well Name",       "count"),
               PoorPerformance_count = ("PoorPerformance", "sum"),
               SpeedUp_count         = ("SpeedUp",         "sum"),
               HighMotorTemp_count   = ("HighMotorTemp",   "sum"),
               MissingSensor_count   = ("MissingSensor",   "sum"),
               ModemOffline_count    = ("ModemOffline",    "sum"),
           )
           .reset_index()
    )
    cust_metrics = cust_metrics[cust_metrics["Customer"].str.strip() != ""]


    # ───────────── Summary Metrics Row (Customer Overview) ─────────────
    num_customers = cust_metrics["Customer"].nunique()
    total_wells   = cust_metrics["WellCount"].sum()
    total_poor    = int(cust_metrics["PoorPerformance_count"].sum())
    total_speedup = int(cust_metrics["SpeedUp_count"].sum())
    total_temp    = int(cust_metrics["HighMotorTemp_count"].sum())
    total_missing = int(cust_metrics["MissingSensor_count"].sum())
    total_modem = int((df3["ModemOffline"]).sum())

    cols = st.columns([2, 1, 1, 1, 1, 1, 1])
# ─── Customer-Summary card (click → show all customers in Dashboard) ──────────
    with cols[0]:
        # two‐line Markdown label
        summary_label = "Customer Summary\n" + f"{num_customers} customers"
        clicked = st.button(
            summary_label,
            key="card_summary",
            help="Click to open the Dashboard with every customer",
            use_container_width=True,
        )

        # style the <button> inside the <div data-testid="card_summary">
        st.markdown("""
        <style>
            /* target the wrapper-div by your key, then its inner <button> */
            div[data-testid="card_summary"] button {
                background-color: #002062 !important;   /* blue bg */
                color:           #f4bb2a !important;   /* gold text */
                border: none !important;
                border-radius: 8px !important;
                box-shadow: 0 2px 4px rgba(0,0,0,0.15) !important;

                /* same height as your metric cards */
                height: 80px !important;
                width:  100% !important;

                /* stack & center the two lines */
                display:        flex !important;
                flex-direction: column !important;
                justify-content: center !important;
                align-items:     center !important;

                padding: 0 !important;  /* rely on height */
            }

            /* “Customer Summary” line */
            div[data-testid="card_summary"] button > div:first-child {
                font-size: 18px !important;
                font-weight: 600 !important;
                line-height: 1 !important;
            }
            /* “XX customers” line */
            div[data-testid="card_summary"] button > div:last-child {
                font-size: 14px !important;
                margin-top: 4px !important;
            }
            </style>
            """, unsafe_allow_html=True)

        if clicked:
            st.session_state.selected_customer = None
            st.session_state.view_page = "Dashboard"
            st.rerun()


    # Total Wells card
    with cols[1]:
        st.markdown(f"""
            <div style="
                background-color:#002062;
                color:#f4bb2a;
                padding:12px;
                border-radius:8px;
                box-shadow:0 2px 4px rgba(0,0,0,0.15);
                text-align:center;
            ">
            <h2 style="margin:0;">Total Wells</h2>
            <p style="margin:4px 0 0; font-size:16px;">
                <strong>{total_wells}</strong>
            </p>
            </div>
        """, unsafe_allow_html=True)

    # helper to render your flag cards
    def render_flag_card(col, title, value, light_bg, dark_bg, light_txt, dark_txt):
        style = (
            f"background-color:{dark_bg};color:{dark_txt};"
            if night_mode else
            f"background-color:{light_bg};color:{light_txt};"
        )
        with col:
            st.markdown(f"""
                <div style="
                    {style}
                    padding:12px;
                    border-radius:8px;
                    text-align:center;
                    box-shadow:0px 2px 4px rgba(0,0,0,0.15);
                ">
                    <div style="font-size:14px;font-weight:600;margin-bottom:4px;">
                        {title}
                    </div>
                    <div style="font-size:24px;font-weight:bold;">
                        {value}
                    </div>
                </div>
            """, unsafe_allow_html=True)

    # Summary flag cards
    render_flag_card(cols[2], "Poor Performance", total_poor,
                     "#FDE0E0", "#8B0000", "#B00000", "#FFFFFF")
    render_flag_card(cols[3], "Speed Up", total_speedup,
                     "#E0FDE0", "#006400", "#006400", "#FFFFFF")
    render_flag_card(cols[4], "High Motor Temp", total_temp,
                     "#FFF1E0", "#CC8400", "#CC6600", "#FFFFFF")
    render_flag_card(cols[5], "Missing Sensor", total_missing,
                     "#F0E0FD", "#6A0080", "#8000CC", "#FFFFFF")
    render_flag_card(cols[6],"Modem Offline",total_modem,
                     "#E0E0E0","#444444","#000000", "#FFFFFF")
    # ───── Render custom cards ─────
    custom = settings.get(current_key, {}).get("custom_cards", [])
    if custom:
        # up to 5 cards in the same layout
        cols_custom = st.columns([2] + [1] * min(len(custom), 5))
        for idx, card in enumerate(custom[:5]):
            with cols_custom[idx + 2]:
                # build & combine each condition
                mask = None
                for cond in card["conditions"]:
                    f, o, v = cond["field"], cond["op"], cond["value"]
                    if   o == ">": m = df3[f] >  v
                    elif o == "<": m = df3[f] <  v
                    else:           m = df3[f] == v

                    if mask is None:
                        mask = m
                    else:
                        if card.get("combiner","AND") == "AND":
                            mask = mask & m
                        else:
                            mask = mask | m

                count = int(mask.sum())
                bg    = card["color"] + "33"
                txt   = card["color"]
                st.markdown(f"""
                    <div style="
                        background-color:{bg};
                        color:{txt};
                        padding:12px;
                        border-radius:8px;
                        text-align:center;
                        box-shadow:0 2px 4px rgba(0,0,0,0.15);
                    ">
                      <div style="font-size:14px;font-weight:600;">
                        {card['label']}
                      </div>
                      <div style="font-size:24px;font-weight:bold;">
                        {count}
                      </div>
                    </div>
                """, unsafe_allow_html=True)

    st.markdown("---")
    # ───────────── Per‐Customer Row Cards ─────────────
    for i, row in cust_metrics.iterrows():
        cust_name = row["Customer"]
        c0, c1, c2, c3, c4, c5, c6 = st.columns([2,1,1,1,1,1,1])
        with c0:
            cust_label = f"**{cust_name}**"
            clicked = st.button(
                cust_label,
                key=f"cust_card_{i}",
                help=f"Open dashboard for {cust_name}",
                use_container_width=True,
            )
            st.markdown(f"""
            <style>
            /* inner button of this particular card */
            div[data-testid="cust_card_{i}"] button {{
                background-color: #002062 !important;
                color:           #f4bb2a !important;
                border: none !important;
                border-radius: 8px !important;
                box-shadow: 0 2px 4px rgba(0,0,0,0.15) !important;

                height: 80px !important;
                width:  100% !important;
                display:        flex !important;
                justify-content: center !important;
                align-items:     center !important;
                padding: 0 !important;
            }}
            div[data-testid="cust_card_{i}"] button > div:first-child {{
                font-size: 18px !important;
                font-weight: 600 !important;
            }}
            </style>
            """, unsafe_allow_html=True)

            if clicked:
                st.session_state.selected_customer = cust_name
                st.session_state.view_page        = "Dashboard"
                st.rerun()
            # ───── CUSTOM SETTINGS NOTICE & RESET — side-by-side ─────
            if cust_name in settings and cust_name != "DEFAULT":
                with st.container():
                    col_notice, col_button = st.columns([3, 2])
                    with col_notice:
                        st.markdown(
                            "<span style='color:#cc0000; font-weight:600;'>"
                            "⚙️ Custom settings in use. Reset to default?"
                            "</span>",
                            unsafe_allow_html=True
                        )
                    with col_button:
                        with st.form(key=f"reset_form_{i}", clear_on_submit=True):
                            submit = st.form_submit_button("Reset to default settings")
                            if submit:
                                settings.pop(cust_name, None)
                                SETTINGS_FILE.write_text(json.dumps(settings, indent=2))
                                st.success(f"Settings for {cust_name} reset to default.")
                                st.rerun()
            # ───── DRIVE TYPE COUNTS ─────
            cust_drives = drive_counts_df[drive_counts_df["Customer"] == cust_name]
            if not cust_drives.empty:
                st.markdown("**Drive types:**")
                for _, dc in cust_drives.iterrows():
                    st.markdown(f"- {dc['Drive Type']}: {dc['Count']}")


        
        # 2) Wells‐count card
        with c1:
            st.markdown(f"""
                <div style="
                    background-color:#002062;
                    color:#f4bb2a;
                    padding:12px;
                    border-radius:8px;
                    box-shadow:0 2px 4px rgba(0,0,0,0.15);
                    text-align:center;
                ">
                <h2 style="margin:0;">{row["WellCount"]}</h2>
                </div>
            """, unsafe_allow_html=True)
        # 3–6) Flags for each customer
        render_flag_card(c2, "Poor Performance", int(row["PoorPerformance_count"]),
                         "#FDE0E0", "#8B0000", "#B00000", "#FFFFFF")
        render_flag_card(c3, "Speed Up", int(row["SpeedUp_count"]),
                         "#E0FDE0", "#006400", "#006400", "#FFFFFF")
        render_flag_card(c4, "High Motor Temp", int(row["HighMotorTemp_count"]),
                         "#FFF1E0", "#CC8400", "#CC6600", "#FFFFFF")
        render_flag_card(c5, "Missing Sensor", int(row["MissingSensor_count"]),
                         "#F0E0FD", "#6A0080", "#8000CC", "#FFFFFF")
        render_flag_card(c6,"Modem Offline",int(row["ModemOffline_count"]),
                     "#E0E0E0","#444444","#000000", "#FFFFFF")
        

    st.stop()

# ───────────── DASHBOARD (per-customer) ─────────────
if page == "Dashboard":
    if st.button("← Back to Customers"):
        st.session_state.selected_customer = None
        st.session_state.view_page = "Customers"
        st.rerun()
        # if a customer has been picked, only keep that subset
    cust = st.session_state.selected_customer
    if cust:
        # ensure df3 has a Customer column
        df3 = df3.copy()
        df3["Customer"] = df3["Well Name"].map(well2cust)
        df3 = df3[df3["Customer"] == cust]

        # same for the AG-grid display frame
        df_show = df_show.copy()
        df_show["Customer"] = df_show["Well Name"].map(well2cust)
        df_show = df_show[df_show["Customer"] == cust]

    # 1) compute the four counts from df3:
    poor_count      = int(df3["PoorPerformance"].sum())
    speedup_count   = int(df3["SpeedUp"].sum())
    hightemp_count  = int(df3["HighMotorTemp"].sum())
    missing_count   = int(df3["MissingSensor"].sum())
    modem_offline_count = int(df3["ModemOffline"].sum())

    # 2) Create a row of 5 columns: one big for title/text, and four small for cards
    col_title, col_poor, col_speedup, col_ht, col_miss, col_modem = st.columns([4,1,1,1,1,1])

    # 2a) Title + “Loaded X wells” goes in the first column
    with col_title:

        # Dashboard title (per-customer or all)
        cust = st.session_state.selected_customer
        title_text = (
            f"{cust} Daily Well-Performance Dashboard"
            if cust else
            "Daily Well-Performance Dashboard"
        )
        st.markdown(
            f"<h1 style='margin:0; text-align:center; "
            f"color:{'#ffffff' if night_mode else '#000000'};'>"
            f"{title_text}</h1>",
            unsafe_allow_html=True,
        )
        # “Loaded X wells from …” text
        date_range = (
            f"{min(last_dates)}"
            if hist_days == 1
            else f"{min(last_dates)} → {max(last_dates)}"
        )
        st.markdown(
            f"<p style='margin:0; font-size:14px; "
            f"color: {'#e0e0e0' if night_mode else '#333'};'>"
            f"Loaded <strong>{len(df3)}</strong> wells from {hist_days} day"
            f"{'s' if hist_days>1 else ''} ({date_range})</p>",
            unsafe_allow_html=True,
        )

    # 2b) Four cards, one per metric:

    # PoorPerformance (red)
    light_bg, dark_bg = "#FDE0E0", "#8B0000"
    light_txt, dark_txt = "#B00000", "#FFFFFF"
    with col_poor:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    Poor Performance
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {poor_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # SpeedUp (green)
    light_bg, dark_bg = "#E0FDE0", "#006400"
    light_txt, dark_txt = "#006400", "#FFFFFF"
    with col_speedup:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    Speed Up
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {speedup_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # HighMotorTemp (orange)
    light_bg, dark_bg = "#FFF1E0", "#CC8400"
    light_txt, dark_txt = "#CC6600", "#FFFFFF"
    with col_ht:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    High Motor Temp
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {hightemp_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # MissingSensor (purple)
    light_bg, dark_bg = "#F0E0FD", "#6A0080"
    light_txt, dark_txt = "#8000CC", "#FFFFFF"
    with col_miss:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    Missing Sensor
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {missing_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    # ─── Modem Offline ─────────────────────────────────────────
    light_bg, dark_bg = "#E0E0E0", "#444444"
    light_txt, dark_txt = "#000000", "#FFFFFF"
    with col_modem:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    Missing Sensor
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {modem_offline_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ───── Render custom cards ─────
    custom = settings.get(current_key, {}).get("custom_cards", [])
    if custom:
        # up to 5 cards, blank spacer to align under “Poor Performance”
        num_cards   = min(len(custom), 5)
        cols_custom = st.columns([4] + [1] * num_cards)

        for idx, card in enumerate(custom[:num_cards]):
            # make sure we have a list of conditions
            if "conditions" not in card:
                continue  # skip any malformed entries
            with cols_custom[idx + 1]:
                mask = None
                for cond in card["conditions"]:
                    f, o, v = cond["field"], cond["op"], cond["value"]
                    if   o == ">": m = df3[f] >  v
                    elif o == "<": m = df3[f] <  v
                    else         : m = df3[f] == v

                    mask = m if mask is None else (
                        (mask & m) if card.get("combiner","AND")=="AND" else (mask | m)
                    )

                count = int(mask.sum()) if mask is not None else 0
                bg, txt = card["color"] + "33", card["color"]

                st.markdown(f"""
                    <div style="
                        background-color:{bg};
                        color:{txt};
                        padding:12px;
                        border-radius:8px;
                        text-align:center;
                        box-shadow:0 2px 4px rgba(0,0,0,0.15);
                    ">
                    <div style="font-size:14px;font-weight:600;">
                        {card['label']}
                    </div>
                    <div style="font-size:24px;font-weight:bold;">
                        {count}
                    </div>
                    </div>
                """, unsafe_allow_html=True)

    st.markdown("---")

    options = ["All"] + list(flag_map.keys())
    flag    = st.selectbox("Filter wells by flag", options)
    view = df_show if flag == "All" else df_show[df3[flag]]
    view = view.sort_values("TerribleScore", ascending=False)

    grid_theme = "ag-theme-alpine-dark" if night_mode else "ag-theme-alpine"
    from st_aggrid import GridUpdateMode, DataReturnMode

    grid_response = AgGrid(
        view,
        gridOptions=grid_opts,
        allow_unsafe_jscode=True,
        theme=grid_theme,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        use_container_width=True,
        fit_columns_on_grid_load=False,
    )
    
    df_live = pd.DataFrame(grid_response["data"])[display_cols]
       # ─── PDF download of the live AG-Grid via wkhtmltopdf ───────────────────
   # Dependencies:
   #   pip install pdfkit
   #   sudo apt-get install -y wkhtmltopdf
   import pdfkit, json

   # Serialize your existing grid options & the current row data
   grid_json = json.dumps(grid_opts)
   row_json  = json.dumps(df_live.to_dict(orient="records"))

   # Build a minimal HTML page (with AG-Grid CSS/JS from CDN)
   html = f"""
   <!DOCTYPE html>
   <html>
   <head>
     <link rel="stylesheet" href="https://unpkg.com/ag-grid-community/dist/styles/ag-grid.css">
     <link rel="stylesheet" href="https://unpkg.com/ag-grid-community/dist/styles/ag-theme-alpine.css">
     <style>html, body, #myGrid {{height:100%;margin:0;padding:0;}}</style>
   </head>
   <body>
     <div id="myGrid" class="{'ag-theme-alpine-dark' if night_mode else 'ag-theme-alpine'}"></div>
     <script src="https://unpkg.com/ag-grid-community/dist/ag-grid-community.min.noStyle.js"></script>
     <script>
       var gridOptions = {grid_json};
       gridOptions.rowData = {row_json};
       new agGrid.Grid(document.getElementById('myGrid'), gridOptions);
     </script>
   </body>
   </html>
   """

   # Convert that HTML → PDF in memory
   pdf_bytes = pdfkit.from_string(html, False, options={'enable-local-file-access': ''})
   st.download_button(
       label="📥 Download current table as PDF",
       data=pdf_bytes,
       file_name=f"well_report_{today}.pdf",
       mime="application/pdf",
   )
    # Excel download of the same
    from io import BytesIO
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        df_live.to_excel(writer, index=False, sheet_name="Wells")
    buf.seek(0)
    st.download_button(
        label="📥 Download current table as Excel",
        data=buf,
        file_name=f"well_report_{today}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


else:
    # ─────────── “Raw Data” tab (unchanged) ───────────
    st.subheader("Full aggregated df3")
    if night_mode:
        st.dataframe(
            df3.style.set_properties(
                **{"background-color": "#2d2d2d", "color": "#e0e0e0", "border-color": "#555"}
            ),
            height=700,
        )
    else:
        st.dataframe(df3, height=700)


# ─── If we have an n8n response, show it here ───────────────────────
if "n8n_response" in st.session_state:
    st.markdown(f"### n8n tickets for “{st.session_state.get('last_triggered_well','')}”")
    df_tickets = pd.DataFrame(st.session_state["n8n_response"])
    st.table(df_tickets)
# To run from PowerShell or Command Prompt:
# cd "C:\\Users\\Thai.phi\\OneDrive - Endurance Lift Solutions\\Desktop\\modbus dashboard"
# cd "C:\\Users\\Thai.phi\\OneDrive - Endurance Lift Solutions\\Desktop\\modbus dashboard"
# streamlit run "well review.v2.py"
