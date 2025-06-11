# well_review.py – Daily Well‐Performance Dashboard (with clickable Well Name, logo, and night mode)
# -----------------------------------------------------------------------------------------
# • Accepts .csv, .xls, .xlsx  (upload or auto‐scan data/)
# • Excel: first sheet, header row=5, report date from AK2
# • For .xlsx: extracts true hyperlinks from “Links” via openpyxl
# • For .xls: reads with pandas/xlrd and sets Link URL = None
# • Aggregates last 3 days, builds flags & Terrible‐Performance score
# • AG‐Grid with pinned “Well Name” & “TerribleScore”
# • Wide layout, resizable columns, color‐coded cells
# • New columns: Running Days, Drive Type, State Detail/Op Mode,
#   Fault Count (cumulative), NearUnderload Ratio & “✗” flag
# • “Normal_vs_Overload” shows a red “✗” (no background)
# • Dashboard <-> Raw‐data page toggle
# • Top‐corner: company logo + contact info
# • Night mode toggle in sidebar

import pathlib, re, datetime as dt, numpy as np, pandas as pd, streamlit as st
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
import openpyxl
import zipfile, xlrd
from io import BytesIO

# ───────────────────────── Page config & theme toggle ─────────────────────────
st.set_page_config(layout="wide")
DATA_DIR      = pathlib.Path("data")
LOOKBACK_DAYS = 4   # today + previous 3 days
ROLL_DAYS     = 3
today         = dt.date.today()
csv_date_re   = re.compile(r"(\d{4}-\d{2}-\d{2})", re.I)


# ──────────── Sidebar: Night Mode Toggle & CSS Overrides ─────────────────
night_mode = st.sidebar.checkbox("Night mode")
if night_mode:
    st.markdown(
        """
        <style>
        /* ─── 1) Overall page + block container ───────────────────────────────── */
        /* Make the main body background nearly black, with light text */
        [data-testid="stAppViewContainer"] {
            background-color: #1e1e1e !important;
            color: #e0e0e0 !important;
        }
        /* If you still see .block-container used, override that too: */
        .block-container {
            background-color: #1e1e1e !important;
            color: #e0e0e0 !important;
        }

        /* ─── 2) Sidebar background, headings, and text ───────────────────────── */
        [data-testid="stSidebar"] {
            background-color: #252526 !important;
            color: #e0e0e0 !important;
        }
        /* Force every piece of text inside the sidebar to be light */
        [data-testid="stSidebar"] * {
            color: #e0e0e0 !important;
        }
        /* Sidebar headers (section titles) */
        .css-1d391kg, .css-1v3fvcr, .css-10trblm {
            /* These class names can vary by Streamlit version. 
               If your headings remain dark, inspect and override them here. */
            color: #e0e0e0 !important;
        }

        /* ─── 3) Main headings & markdown text ─────────────────────────────── */
        h1, h2, h3, h4, h5, h6 {
            color: #ffffff !important;
        }
        /* Standard markdown text */
        .css-901oao, .css-16huue1, .css-1n76uvr {
            /* These classes are for normal paragraph text. */
            color: #e0e0e0 !important;
        }
        /* Links inside markdown */
        a {
            color: #66CDAA !important;
        }

        /* ─── 4) Streamlit widgets: buttons, inputs, selects, checkboxes ───────── */
        /* Buttons */
        .stButton>button {
            background-color: #3a3d41 !important;
            color: #ffffff !important;
            border: 1px solid #555 !important;
        }
        .stButton>button:hover {
            background-color: #50545a !important;
        }
        /* Text input boxes and number_input boxes */
        input[type="text"], input[type="number"] {
            background-color: #2d2d2d !important;
            color: #e0e0e0 !important;
            border: 1px solid #555 !important;
        }
        /* Selectboxes, Radio buttons, Checkboxes */
        .stSelectbox, .stRadio, .stCheckbox {
            background-color: #2d2d2d !important;
            color: #e0e0e0 !important;
        }
        /* The actual <label> text for slider, checkbox, etc. */
        label, .css-1v0mbdj, .css-1f4mp12 {
            color: #e0e0e0 !important;
        }
        /* Slider track & handle */
        div[data-baseweb="slider"] {
            background-color: #2d2d2d !important;
        }
        div[data-baseweb="slider"] .rc-slider-rail {
            background-color: #555 !important;
        }
        div[data-baseweb="slider"] .rc-slider-track {
            background-color: #66CDAA !important;
        }
        div[data-baseweb="slider"] .rc-slider-handle {
            background-color: #66CDAA !important;
            border: 1px solid #444 !important;
        }

        /* ─── 5) DataFrame (st.dataframe) override ────────────────────────────── */
        .stDataFrame table {
            background-color: #2d2d2d !important;
            color: #e0e0e0 !important;
        }
        .stDataFrame th {
            background-color: #333 !important;
            color: #e0e0e0 !important;
        }
        .stDataFrame td {
            border-color: #555 !important;
        }

        /* ─── 6) AG-Grid container & cells ───────────────────────────────────── */
        /* (a) Header row: keep dark gray + white text */
        .ag-theme-alpine-dark .ag-header,
        .ag-theme-alpine-dark .ag-header-cell,
        .ag-theme-alpine-dark .ag-header-row {
            background-color: #333 !important;
            color: #e0e0e0 !important;
            border-bottom: 1px solid #555 !important;
        }

        /* (b) Default all data-row cells to white + black text.
            Any inline `js_color` background (green/red) will override this white where it applies. */
        .ag-theme-alpine-dark .ag-center-cols-container .ag-row .ag-cell {
            background-color: #ffffff;    /* plain white */
            color: #000000;               /* plain black */
        }

        /* (c) “Well Name” column cells: force them back to a dark background
            so that your hyperlink‐color logic (blue/red) stands out on dark. */
        .ag-theme-alpine-dark .ag-center-cols-container .ag-cell[col-id="Well Name"] {
            background-color: #2d2d2d !important;
            /* Note: we are NOT forcing a text color here, so your existing
            JS‐renderer (blue link or red “no link”) will still apply. */
        }

        /* (d) Inside “Well Name”, we let <a> tags keep whatever color they already had.
            If you need a custom link‐color, you can add it here. */
        .ag-theme-alpine-dark .ag-center-cols-container .ag-cell[col-id="Well Name"] a {
            /* do NOT override color—your JS renderer’s <a> color stays. */
        }

        /* (e) If your grid has filter/sort icons in the header, brighten them so they remain visible. */
        .ag-theme-alpine-dark .ag-icon {
            filter: brightness(200%) !important;
        }

        /* (f) Ensure the grid’s wrapper behind the header is dark,
            so you don’t see any stray white flashes around the edges. */
        .ag-root-wrapper {
            background-color: #2d2d2d !important;
        }

        /* (g) (Optional) Make the scroll‐thumb a bit darker to blend with night mode */
        .ag-theme-alpine-dark ::-webkit-scrollbar-thumb {
            background-color: #888 !important;
        }
        // ─── Streamlit File Uploader ───────────────────────────── */
        [data-testid="stFileUploader"] {
            background-color: transparent !important;
            color: #ffffff !important;
        }

        [data-testid="stFileUploader"] section {
            background-color: #2d2d2d !important;
            border: 1px solid #555 !important;
            border-radius: 8px !important;
        }

        [data-testid="stFileUploader"] section * {
            color: #ffffff !important;
        }

        [data-testid="stFileDropzone"] {
            background-color: #2d2d2d !important;
            border: 2px dashed #888 !important;
            color: #ffffff !important;
        }

        [data-testid="stFileDropzone"] * {
            color: #ffffff !important;
        }
        [data-testid="stFileUploader"] button {
            background-color: #3a3d41 !important;
            color: #ffffff !important;
            border: 1px solid #777 !important;
            border-radius: 6px !important;
        }

        [data-testid="stFileUploader"] button:hover {
            background-color: #50545a !important;
            color: #ffffff !important;
        }
        /* ─── 8) Scrollbars ───────────────────────────────────────────────────── */
        /* Optional: style scrollbars dark */
        ::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }
        ::-webkit-scrollbar-track {
            background: #1e1e1e;
        }
        ::-webkit-scrollbar-thumb {
            background-color: #555;
            border-radius: 4px;
        }
        ::-webkit-scrollbar-thumb:hover {
            background-color: #666;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

# ───────────────────────── Header: Logo & Contact Info ─────────────────────────
col1, col2 = st.columns([1, 4])
with col1:
    # Make sure "company_logo.png" exists in the same folder, or change the filename here
    st.image("company_logo.png", width=120)
with col2:
    st.markdown(
        """
        <p style="font-size:16px; margin-bottom:4px;">
            <strong>Contact:</strong>
            <a href="mailto:optimization@endurancelift.com" style="color:#0066cc;">
                optimization@endurancelift.com
            </a>
        </p>
        <p style="font-size:16px; margin-top:0;">
            <strong>24 hr hotline:</strong> +1 432 224 0583
        </p>
        """,
        unsafe_allow_html=True,
    )

# ───────────── Helper: ensure Date column ─────────────
def ensure_date_column(df: pd.DataFrame, source_name: str, *, excel_date=None):
    if "Date" not in df.columns:
        if excel_date is not None:
            df.insert(0, "Date", pd.to_datetime(excel_date).date())
        else:
            m = csv_date_re.search(source_name)
            day = m.group(1) if m else today.strftime("%Y-%m-%d")
            df.insert(0, "Date", pd.to_datetime(day).date())
    return df

# ───────────── Excel / CSV loaders ─────────────

def load_excel(buf):
    """
    Return a DataFrame with a Date column (from AK2)
    and a 'Link URL' column (real hyperlink or None).

    • .xlsx/.xlsm  → openpyxl for hyperlinks.
    • .xls         → xlrd 1.2 for data + hyperlinks (no pandas engine).
    """
    raw_bytes = buf.read()
    fname     = buf.name.lower()

    # ---------- .xlsx / .xlsm ------------------------------------------
    if fname.endswith((".xlsx", ".xlsm")):
        # 1) Read visible data via pandas (header row = 5 → index=4)
        bio = BytesIO(raw_bytes)
        df  = pd.read_excel(bio, sheet_name=0, header=4)

        # 2) Extract “report date” from AK2
        bio.seek(0)
        excel_dt = pd.read_excel(
            bio, sheet_name=0, header=None, usecols="AK", nrows=2
        ).iloc[1, 0]
        df = ensure_date_column(df, fname, excel_date=excel_dt)

        # 3) Default “Link URL” = None
        df["Link URL"] = None

        # 4) Attempt hyperlink extraction with openpyxl
        try:
            wb = openpyxl.load_workbook(
                BytesIO(raw_bytes), read_only=True, data_only=True
            )
            ws = wb.active

            # 4a) Find any header cell in row 5 whose text (lowercased) contains "link"
            link_col = None
            for c in range(1, ws.max_column + 1):
                header_val = ws.cell(row=5, column=c).value
                if isinstance(header_val, str) and "link" in header_val.strip().lower():
                    link_col = c
                    break

            if link_col is not None:
                # 4b) For each data row i (Excel row 6 + i), grab hyperlink.target if present
                links = []
                for i in range(len(df)):  # i = 0..len(df)-1 → Excel row = 6+i
                    cell = ws.cell(row=6 + i, column=link_col)
                    links.append(cell.hyperlink.target if cell.hyperlink else None)
                df["Link URL"] = links

        except (zipfile.BadZipFile, openpyxl.utils.exceptions.InvalidFileException):
            pass  # If loading fails, leave Link URL as all None

        return df

    # ---------- .xls ----------------------------------------------------
    if fname.endswith(".xls"):
        # Ensure xlrd 1.2.0 (the last version supporting .xls hyperlinks)
        if xlrd.__version__ != "1.2.0":
            raise RuntimeError(
                f"Your xlrd version is {xlrd.__version__}. Please install xlrd==1.2.0:\n"
                "    pip install xlrd==1.2.0"
            )

        book  = xlrd.open_workbook(file_contents=raw_bytes, formatting_info=True)
        sheet = book.sheet_by_index(0)

        # 1) Read header row (Excel row 5 → index=4)
        header_vals = sheet.row_values(4)
        cols = [str(v).strip() for v in header_vals]

        # 2) Pull all data rows (Excel rows 6 .. end)
        data_rows = [sheet.row_values(r) for r in range(5, sheet.nrows)]
        df = pd.DataFrame(data_rows, columns=cols)

        # 3) Extract “report date” from AK2 (row 2, col AK = index 36)
        excel_dt = sheet.cell_value(1, 36)
        df = ensure_date_column(df, fname, excel_date=excel_dt)

        # 4) Default “Link URL” = None
        df["Link URL"] = None

        # 5) Find any header that contains "link"
        link_col = None
        for idx, h in enumerate(cols):
            if isinstance(h, str) and "link" in h.strip().lower():
                link_col = idx
                break

        # 6) If hyperlink_map exists, extract url_or_path for each data row
        if link_col is not None and hasattr(sheet, "hyperlink_map"):
            links = []
            for r in range(5, sheet.nrows):  # r = 5 .. n-1 → Excel row = r+1
                hlink = sheet.hyperlink_map.get((r, link_col))
                links.append(hlink.url_or_path if hlink else None)
            df["Link URL"] = links

        return df

    # ---------- any other extension (shouldn’t happen) ------------------
    raise ValueError(f"Unsupported Excel file type: {fname}")

def load_csv(buf, name) -> pd.DataFrame:
    df = pd.read_csv(buf)
    df = ensure_date_column(df, name)
    df["Link URL"] = None
    return df

# ───────────── File input (upload / fallback) ─────────────
st.sidebar.header("CSV / Excel Source")
upl = st.sidebar.file_uploader(
    "⬆️ Upload 1–4 files (.csv, .xls, .xlsx)",
    type=["csv", "xls", "xlsx"], accept_multiple_files=True
)

dfs = []
if upl:
    for uf in upl:
        if uf.name.lower().endswith(".csv"):
            dfs.append(load_csv(uf, uf.name))
        else:
            dfs.append(load_excel(uf))
else:
    files = [f for f in DATA_DIR.glob("*") if f.suffix.lower() in {".csv", ".xls", ".xlsx"}]
    recent = [
        f for f in files
        if (today - dt.date.fromtimestamp(f.stat().st_mtime)).days < LOOKBACK_DAYS
    ]
    if not recent:
        st.error("❌ No recent files found and none uploaded.")
        st.stop()
    for f in sorted(recent, key=lambda x: x.stat().st_mtime):
        if f.suffix.lower() == ".csv":
            dfs.append(load_csv(f, f.name))
        else:
            dfs.append(load_excel(f))

df_raw = pd.concat(dfs, ignore_index=True)

# 1) Drop columns whose name is empty or all whitespace:
blank_cols = [c for c in df_raw.columns if not str(c).strip()]
df_raw = df_raw.drop(columns=blank_cols, errors="ignore")

# 2) If any column names were duplicated, keep only the first occurrence:
df_raw = df_raw.loc[:, ~df_raw.columns.duplicated()]

# 3) Now filter out empty Well Names and coerce Date → date
df_raw = (
    df_raw
      .dropna(subset=["Well Name"])
      .loc[lambda d: d["Well Name"].astype(str).str.strip().ne("")]
      .assign(Date=lambda d: pd.to_datetime(d["Date"]).dt.date)
)


# 🔍 DEBUG: Show first 5 uptime rows if requested
if st.sidebar.checkbox("DEBUG: show first 5 uptime rows"):
    st.write(df_raw[["Well Name", "Uptime (%)"]].head())

# ───────────── Sidebar thresholds & weights ─────────────
st.sidebar.header("Thresholds")
thr = dict(
    CapLoadPct     = st.sidebar.slider(
        "Cap load pct (Max / Normal)", 0.50, 5.0, 1.05, 0.01,
        help=(
            "CapLoad = (Max Drive Amps) ÷ (Normal Running Amps).\n"
            "• We flag At_Max_Capacity whenever CapLoad ≥ this slider’s value."
            "• We also use this same threshold to decide if there’s “spare load” in SpeedUp logic."
        )
    ),
    RiskPct        = st.sidebar.slider(
        "Risk pct (Max / Overload)", 0.50, 5.0, 1.05, 0.01,
        help=(
            "CapRisk = (Max Drive Amps) ÷ (Motor Overload).\n"
            "We flag Overload_Risk whenever CapRisk ≥ this slider’s value."
        )
    ),
    HighIntake     = st.sidebar.number_input(
        "High intake ψ (psi)", 0, 10000, 300,
        help=(
            "High intake threshold is used in SpeedUp logic.\n"
            "We require Avg Intake Pressure > [this value] to consider speeding up."
        )
    ),
    SmallDrawdown  = st.sidebar.number_input(
        "Small drawdown ψ (psi)", 0, 5000, 50,
        help=(
            "Drawdown = (Max Intake Pressure – Min Intake Pressure).\n"
            "We require Drawdown < [this value] to consider speeding up."
        )
    ),

    NearUnderLower = st.sidebar.slider(
        "Near-underload lower bound", 1.0, 5.0, 1.43, 0.01,
        help=(
            "NearUnderload Ratio = (Avg Drive Amps) ÷ (Motor Underload).\n"
            "We flag NearUnderload when ratio < [this slider’s value]."
        )
    ),
    LowUptime      = st.sidebar.slider(
        "Low uptime % threshold", 0, 100, 90,
        help=(
            "If (Uptime %) < [this value], we flag LowUptime.\n"
            "Used in TerribleScore and PoorPerformance."
        )
    ),
    HighDT         = st.sidebar.number_input(
        "High downtime hrs (3 days)", 0, 72, 6,
        help=(
            "Downtime (Hr) = hours of downtime over the lookback window.\n"
            "If Downtime > [this value], we flag HighDT (for coloring only)."
        )
    ),
    VibHigh        = st.sidebar.number_input(
        "High vibration threshold", 0.0, 10.0, 0.70, 0.01,
        help=(
            "If (Avg Vib X ≥ [this]) or (Avg Vib Y ≥ [this]), we flag HighVib.\n"
            "Used in TerribleScore and PoorPerformance."
        )
    ),
    SpreadRatio    = st.sidebar.number_input(
        "Spread ratio ≥ (unitless)", 0.0, 10.0, 1.0, 0.01,
        help=(
            "SpreadRatio = (Max Drive Amps – Min Drive Amps) ÷ (Avg Drive Amps).\n"
            "If Min Drive Amps = 0, we display ‘N/A’ and never flag SpreadFlag.\n"
            "We set SpreadFlag when SpreadRatio ≥ [this value]."
        )
    ),
    TempHigh       = st.sidebar.number_input(
        "High motor temp °F", 0, 500, 210,
        help=(
            "If (Max Motor Temp) ≥ [this value], we flag HighMotorTemp.\n"
            "Used in TerribleScore and PoorPerformance."
        )
    ),
    LowDelta       = st.sidebar.number_input(
        "Low Tub-Casing Δ ψ", 0, 5000, 30,
        help=(
            "Δ = (Avg Tubing Pressure) – (Avg Casing Pressure).\n"
            "If Δ ≤ [this value], we flag LowDeltaTC (for coloring only)."
        )
    ),
    HighFaultCount = st.sidebar.number_input(
        "High fault count (cumulative)", 0, 1000, 1,
        help=(
            "If (Fault Count) ≥ [this value], we flag FaultHigh.\n"
            "Used in TerribleScore and PoorPerformance."
        )
    ),
)
st.sidebar.markdown("---")
st.sidebar.subheader("Weights – Terrible‐Performance Score")
weights = dict(
    uptime        = st.sidebar.slider(
        "Downtime weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Adds (LowUptime × [this value]) to TerribleScore.\n"
            "LowUptime = 1 if Uptime% < Low uptime threshold, else 0."
        )
    ),
    missing       = st.sidebar.slider(
        "Missing sensor weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Adds (MissingSensor × [this value]) to TerribleScore.\n"
            "MissingSensor = 1 if Avg Motor Amps=0 or Avg Intake Pressure=0 or flat-line in either; else 0."
        )
    ),
    spread        = st.sidebar.slider(
        "Spread ratio weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Adds (SpreadRatio × [this value]) to TerribleScore.\n"
            "SpreadRatio = (Max Drive Amps − Min Drive Amps) ÷ (Avg Drive Amps)."
        )
    ),
    motortemp     = st.sidebar.slider(
        "Motor temp weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Adds (HighMotorTemp × [this value]) to TerribleScore.\n"
            "HighMotorTemp = 1 if Max Motor Temp ≥ High motor temp threshold, else 0."
        )
    ),
    drawdown      = st.sidebar.slider(
        "Intake drawdown weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Adds ([Drawdown < SmallDrawdown] × [this value]) to TerribleScore.\n"
            "Drawdown = Max Intake Pressure − Min Intake Pressure."
        )
    ),
    delta_tc      = st.sidebar.slider(
        "Tub-Cas Δ weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Removed from TerribleScore (set to 0); kept here only for coloring."
        )
    ),
    nearunderload = st.sidebar.slider(
        "NearUnderload weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Adds (NearUnderload × [this value]) to TerribleScore.\n"
            "NearUnderload = 1 if (Avg Drive Amps ÷ Motor Underload) < Near-underload threshold, else 0."
        )
    ),
    vibration     = st.sidebar.slider(
        "Vibration weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Adds (HighVib × [this value]) to TerribleScore.\n"
            "HighVib = 1 if (Max Vibration ≥ High vibration threshold)."
        )
    ),
    fault         = st.sidebar.slider(
        "Fault count weight", 0.0, 5.0, 1.0, 0.1,
        help=(
            "Adds (FaultHigh × [this value]) to TerribleScore.\n"
            "FaultHigh = 1 if Fault Count ≥ High fault count threshold, else 0."
        )
    ),
)

# ───────────── Aggregate last 3 distinct days per well ─────────────
last_dates = sorted(df_raw["Date"].unique())[-ROLL_DAYS:]
df_recent  = df_raw[df_raw["Date"].isin(last_dates)].copy()
hist_days  = len(last_dates)
use_flat   = hist_days >= 3

# ── RIGHT BEFORE the pd.to_numeric loop: Sanitize Uptime (%) ──
if "Uptime (%)" in df_recent.columns:
    col = (
        df_recent["Uptime (%)"]
            .astype(str)
            .str.strip()
            .str.replace('%', '', regex=False)
            .str.lower()
            .replace({'': np.nan, 'n/a': np.nan, 'na': np.nan, 'nan': np.nan, '--': np.nan})
    )
    num = pd.to_numeric(col, errors='coerce')
    # If ≤1.05 assume fraction (0–1); if >1.05 assume percent (0–100)
    df_recent["Uptime (%)"] = np.where(num <= 1.05, num, num / 100.0)

# Convert all other columns (except text fields) to numeric
non_txt = {
    "Well Name", "Field", "Installation Date", "Current Status",
    "Pump Type", "Drive Type", "State Detail/Op Mode", "Links",
    "Latest Fault", "Fault Date", "Link URL"
}

for c in df_recent.columns:
    if c not in non_txt and c != "Date":
        col_data = df_recent[c]
        if not (isinstance(col_data, pd.Series) and col_data.ndim == 1):
            print(f"Column {c} is not a Series or is not 1D, got type {type(col_data)} and ndim {getattr(col_data,'ndim',None)}")
        else:
            df_recent[c] = pd.to_numeric(col_data, errors="coerce")

numeric_cols = df_recent.select_dtypes(include="number").columns.tolist()
text_cols    = [c for c in df_recent.columns if c not in numeric_cols + ["Date"]]

agg_dict = {
    c: ["mean","max","min","std"]
    for c in numeric_cols
    if c!="Well Name"
}
# wrap the "first" in a list so pandas will do SeriesGroupBy.first, not DataFrameGroupBy.first
agg_dict.update({
    c: ["first"]
    for c in text_cols
    if c!="Well Name"
})
df3 = df_recent.groupby("Well Name").agg(agg_dict)
df3.columns = ["_".join(c) if isinstance(c, tuple) else c for c in df3.columns]
df3 = df3.reset_index()

def col(base: str, stat: str) -> str:
    return f"{base}_{stat}"

# ───────────── Latest-day Normal vs Overload ─────────────
latest = (
    df_raw.sort_values("Date")
          .groupby("Well Name", as_index=False)
          .tail(1)
          .set_index("Well Name")
)
df3["Latest_Normal"]   = latest["Normal Running Amps"].reindex(df3["Well Name"]).values
df3["Latest_Overload"] = latest["Motor Overload"].reindex(df3["Well Name"]).values
df3["Normal_vs_Overload"] = df3["Latest_Normal"] >= df3["Latest_Overload"]

# ───────────── Cap-load & risk (flipped) ─────────────
df3["CapLoad"] = df3[col("Max Drive Amps", "mean")] / df3[col("Normal Running Amps", "mean")]
df3["CapRisk"] = df3[col("Max Drive Amps", "mean")] / df3[col("Motor Overload", "mean")]

df3["At_Max_Capacity"] = df3["CapLoad"] >= thr["CapLoadPct"]
df3["Overload_Risk"]   = df3["CapRisk"] >= thr["RiskPct"]

# ───────────── Additional flags and derived columns ─────────────
drawdown = df3[col("Max Intake Pressure", "max")] - df3[col("Min Intake Pressure", "min")]

df3["NearUnderload Ratio"] = df3[col("Avg Drive Amps", "mean")] / df3[col("Motor Underload", "mean")]
df3["NearUnderload"] = df3["NearUnderload Ratio"] < thr["NearUnderLower"]

df3["Max Vibration"] = df3[[col("Avg Vib X", "mean"), col("Avg Vib Y", "mean")]].max(axis=1)
df3["HighVib"] = df3["Max Vibration"] >= thr["VibHigh"]

df3["HighMotorTemp"] = df3[col("Max Motor Temp", "max")] >= thr["TempHigh"]

df3["Lost_Motor"] = (
    (df3[col("Avg Motor Amps", "mean")] == 0) |
    (use_flat & (df3[col("Avg Motor Amps", "std")] == 0))
)
df3["Lost_Intake"] = (
    (df3[col("Avg Intake Pressure", "mean")] == 0) |
    (use_flat & (df3[col("Avg Intake Pressure", "std")] == 0))
)
df3["MissingSensor"] = df3[["Lost_Motor", "Lost_Intake"]].any(axis=1)

# Uptime %: scale fraction (0–1) → 0–100
uptime_pct       = df3[col("Uptime (%)", "mean")] * 100
df3["LowUptime"] = uptime_pct < thr["LowUptime"]


# ───────────── Compute Spread Ratio & Flag ─────────────
spread_ratio = (
    (df3[col("Max Drive Amps", "max")] - df3[col("Min Drive Amps", "min")])
    / df3[col("Avg Drive Amps", "mean")]
)
df3["SpreadRatio"] = spread_ratio.where(df3[col("Min Drive Amps", "min")] != 0, np.nan)
df3["SpreadFlag"] = (
    (df3["SpreadRatio"] >= thr["SpreadRatio"]) &
    (df3[col("Min Drive Amps", "min")] != 0)
)

# ───────────── Fault Count & Flag ─────────────
df3["Fault Count"] = (df3[col("Fault Count (24hr)", "mean")] * hist_days).round().astype(int)
df3["FaultHigh"]   = df3["Fault Count"] >= thr["HighFaultCount"]

# ───────────── SpeedUp (restored) ─────────────
df3["SpeedUp"] = (
    (df3[col("Running Days", "mean")] < 90) &
    (df3[col("Avg Intake Pressure", "mean")] > thr["HighIntake"]) &
    (drawdown < thr["SmallDrawdown"]) &
    (~df3["Overload_Risk"]) &
    (~df3["At_Max_Capacity"])
)

# ───────────── PoorPerformance (updated) ─────────────
df3["PoorPerformance"] = df3[
    [
        "LowUptime",
        "HighVib",
        "SpreadFlag",
        "HighMotorTemp",
        "FaultHigh",
    ]
].any(axis=1)

# ───────────── Terrible‐Performance Score ─────────────
df3["TerribleScore"] = (
      weights["uptime"]       * df3["LowUptime"].astype(int)
    + weights["missing"]      * df3["MissingSensor"].astype(int)
    + weights["spread"]       * df3["SpreadRatio"].fillna(0)
    + weights["motortemp"]    * df3["HighMotorTemp"].astype(int)
    + weights["drawdown"]     * (drawdown < thr["SmallDrawdown"]).astype(int)
    + weights["nearunderload"]* df3["NearUnderload"].astype(int)
    + weights["vibration"]    * df3["HighVib"].astype(int)
    + weights["fault"]        * df3["FaultHigh"].astype(int)
)

# ───────────── Build AG‐Grid table ───────────────────────────────────────
df_show = df3.copy()
# ─── Add hidden Boolean columns for PoorPerformance reasons ───
df_show["LowUptime_bool"]    = df3["LowUptime"]       # True = failed uptime threshold
df_show["HighVib_bool"]      = df3["HighVib"]         # True = failed vibration threshold
df_show["SpreadFlag_bool"]   = df3["SpreadFlag"]      # True = failed spread‐ratio threshold
df_show["HighMotorTemp_bool"]= df3["HighMotorTemp"]   # True = failed motor‐temp threshold
df_show["FaultHigh_bool"]    = df3["FaultHigh"]       # True = failed fault‐count threshold

# ─── Add hidden Boolean columns for SpeedUp reasons ───
#  1) Running Days < 90?
df_show["Speed_RunDays_OK"]     = df3[col("Running Days", "mean")] < 90

#  2) Avg Intake Pressure > HighIntake?
df_show["Speed_AvgIntake_OK"]   = df3[col("Avg Intake Pressure", "mean")] > thr["HighIntake"]

#  3) Drawdown < SmallDrawdown?
#     (we already computed `drawdown = Max Intake – Min Intake` above)
df_show["Speed_Drawdown_OK"]    = drawdown < thr["SmallDrawdown"]

#  4) Overload_Risk == False?
df_show["Speed_OverloadOK"]     = ~df3["Overload_Risk"].astype(bool)

#  5) At_Max_Capacity == False?
df_show["Speed_AtMaxOK"]        = ~df3["At_Max_Capacity"].astype(bool)

# Numeric columns first
df_show["High Motor Temp"]     = df3[col("Max Motor Temp", "max")]
df_show["High Downtime"]       = df3[col("Downtime (Hr)", "mean")]
df_show["Max Vibration"]       = df3["Max Vibration"]

# Spread ratio / Tubing-Casing Δ / NearUnderload Ratio
df_show["Spread Ratio"]        = df3["SpreadRatio"]
df_show["Tubing-Casing Δ"]     = df3[col("Avg Tubing", "mean")] - df3[col("Avg Casing", "mean")]
df_show["NearUnderload Ratio"] = df3["NearUnderload Ratio"]

# Boolean‐derived columns replaced with numeric or checkmarks:
df_show["NearUnderload"]       = df3["NearUnderload"].apply(lambda x: "✗" if x else "")
df_show["Normal_vs_Overload"]  = df3["Normal_vs_Overload"].apply(lambda x: "✗" if x else "")
df_show["MissingSensor"]       = df3["MissingSensor"].apply(lambda x: "✗" if x else "")

df_show["Fault Count"] = df3["Fault Count"]       # numeric, colored below
df_show["HighVib"]     = df3["Max Vibration"]     # numeric, colored below

df_show["Uptime %"]    = uptime_pct               # numeric, colored below
df_show["PoorPerformance"] = df3["PoorPerformance"].apply(lambda x: "✗" if x else "✓")
df_show["SpeedUp"] = df3["SpeedUp"].apply(lambda x: "✓" if x else "✗")

df_show["At_Max_Capacity"] = df3["CapLoad"]
df_show["Overload_Risk"]   = df3["CapRisk"]

df_show["Running Days"]        = df3[col("Running Days", "mean")]
df_show["Drive Type"]          = df3[col("Drive Type", "first")]
df_show["State Detail/Op Mode"]= df3[col("State Detail/Op Mode", "first")]

# Link URL (first value for each well)
df3["Link URL"] = df3[col("Link URL", "first")]
df_show["Link URL"] = df3["Link URL"]

# Columns to display, in order:
display_cols = [
    "Well Name","Running Days", "TerribleScore", "PoorPerformance", "SpeedUp",
    "At_Max_Capacity", "Overload_Risk",
    "High Motor Temp", "High Downtime", "Max Vibration",
    "Spread Ratio", "Tubing-Casing Δ","Fault Count",
    "HighVib", "Uptime %", "NearUnderload Ratio",
    "NearUnderload", "Normal_vs_Overload",
    "MissingSensor", "Drive Type", "State Detail/Op Mode"
]

# ---------------------- JS Cell‐Style + Link‐Renderer ----------------------
js_color = JsCode(f"""
function(p) {{
    // Convert Python night_mode (True/False) → JS boolean (true/false)
    var isDark = {str(night_mode).lower()};  

    // Define color pairs for light vs dark mode
    var lightRed  = '#FFB3B3';
    var lightGreen = '#C6F7C6';
    var darkRed   = '#E74C3C';    // darker red for contrast
    var darkGreen = '#2ECC71';    // emerald green in dark mode
    var cellText  = isDark ? '#FFFFFF' : '#000000';

    // At_Max_Capacity coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'At_Max_Capacity') {{
        var bg = (p.value >= {thr["CapLoadPct"]}) 
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Overload_Risk coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'Overload_Risk') {{
        var bg = (p.value >= {thr["RiskPct"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // High Motor Temp coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'High Motor Temp') {{
        var bg = (p.value !== null && p.value >= {thr["TempHigh"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // High Downtime coloring: red > threshold, green otherwise
    if (p.colDef.field === 'High Downtime') {{
        var bg = (p.value > {thr["HighDT"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Max Vibration coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'Max Vibration') {{
        var bg = (p.value >= {thr["VibHigh"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Spread Ratio coloring: red ≥ threshold, green otherwise
    if (p.colDef.field === 'Spread Ratio') {{
        var bg = ((p.value !== null) && (p.value >= {thr["SpreadRatio"]}))
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Tubing-Casing Δ coloring: red ≤ threshold, green otherwise
    if (p.colDef.field === 'Tubing-Casing Δ') {{
        var bg = (p.value <= {thr["LowDelta"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // NearUnderload Ratio coloring: red < threshold, green otherwise
    if (p.colDef.field === 'NearUnderload Ratio') {{
        var bg = (p.value < {thr["NearUnderLower"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // NearUnderload boolean: red “✗” if True
    if (p.colDef.field === 'NearUnderload') {{
        if (p.value === '✗') {{
            return {{ 'color': (isDark ? '#FF6961' : '#FF0000') }};
        }}
        return null;
    }}

    // Normal_vs_Overload: red “✗” if True
    if (p.colDef.field === 'Normal_vs_Overload') {{
        if (p.value === '✗') {{
            return {{ 'color': (isDark ? '#FF6961' : '#FF0000') }};
        }}
        return null;
    }}

    // MissingSensor boolean: red “✗” if True
    if (p.colDef.field === 'MissingSensor') {{
        if (p.value === '✗') {{
            return {{ 'color': (isDark ? '#FF6961' : '#FF0000') }};
        }}
        return null;
    }}

    // Fault Count: red if ≥ threshold, green otherwise
    if (p.colDef.field === 'Fault Count') {{
        var bg = (p.value >= {thr["HighFaultCount"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // HighVib: red if ≥ threshold, green otherwise
    if (p.colDef.field === 'HighVib') {{
        var bg = (p.value >= {thr["VibHigh"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // Uptime %: red if < LowUptime threshold, green otherwise
    if (p.colDef.field === 'Uptime %') {{
        var bg = (p.value < {thr["LowUptime"]})
            ? (isDark ? darkRed : lightRed)
            : (isDark ? darkGreen : lightGreen);
        return {{ 'backgroundColor': bg, 'color': cellText }};
    }}

    // PoorPerformance boolean: red “✗” if True (no background)
    if (p.colDef.field === 'PoorPerformance') {{
        if (p.value === '✗') {{
            return {{ 'color': (isDark ? '#FF6961' : '#FF0000') }};
        }}
        return null;
    }}

    // SpeedUp: green “✓” if True (no background)
    if (p.colDef.field === 'SpeedUp') {{
        if (p.value === '✓') {{
            return {{ 'color': (isDark ? '#77DD77' : '#00AA00') }};
        }}
        return null;
    }}

    // If no conditions matched, return null so AG-Grid uses default
    return null;
}}
""")

gb = GridOptionsBuilder.from_dataframe(df_show[display_cols])

# Configure “Well Name” as a clickable hyperlink or red if no URL
gb.configure_column(
    "Well Name",
    pinned="left",
    cellRenderer=JsCode("""
        function (params) {
            const url   = params.data["Link URL"];
            const value = params.value || "";
            if (url) {
                // If there’s a URL, wrap the well name in an <a> tag:
                params.eGridCell.innerHTML =
                    `<a href="${url}" target="_blank">${value}</a>`;
            } else {
                // Otherwise, show the well name in red
                params.eGridCell.innerHTML =
                    `<span style="color:red">${value}</span>`;
            }
            return null;
        }
    """)
)
# Pin “Running Days” so it sits to the left of “Well Name”
gb.configure_column(
    "Running Days",
    pinned="left",
    type=["numericColumn"],
    valueFormatter="x.toFixed(1)",
    headerTooltip="Mean Running Days over lookback window"
)

# Pin TerribleScore with tooltip
gb.configure_column(
    "TerribleScore",
    pinned="left",
    type=["numericColumn"],
    headerTooltip=(
        "TerribleScore = "
        "(LowUptime × uptime weight) + "
        "(MissingSensor × missing weight) + "
        "(SpreadRatio × spread weight) + "
        "(HighMotorTemp × motortemp weight) + "
        "((Max Intake – Min Intake) < SmallDrawdown threshold × drawdown weight) + "
        "(NearUnderload × nearunderload weight) + "
        "(HighVib × vibration weight) + "
        "(FaultHigh × fault count weight)."
    )
)
gb.configure_column(
    "PoorPerformance",
    pinned="left",
    type=["textColumn"],
    headerTooltip=(
        "True if any of {LowUptime, HighVib, SpreadFlag, HighMotorTemp, FaultHigh} is True.\n"
        "Displays a red “✗” if True, blank if False."
    ),
    tooltipValueGetter=JsCode("""
        function(params) {
            // Only show tooltip if the cell value is “✗”
            if (params.value !== '✗') {
                return null;
            }
            var d = params.data;
            var lines = [];
            // Define each sub‐criterion and check its hidden Boolean
            var crits = [
                { name: "LowUptime",     ok: !d.LowUptime_bool    },
                { name: "HighVib",       ok: !d.HighVib_bool       },
                { name: "SpreadFlag",    ok: !d.SpreadFlag_bool    },
                { name: "HighMotorTemp", ok: !d.HighMotorTemp_bool },
                { name: "FaultHigh",     ok: !d.FaultHigh_bool     }
            ];
            crits.forEach(function(c) {
                // If ok===false, that means it “failed”
                var status = c.ok ? "passed" : "failed";
                lines.push(c.name + ": " + status);
            });
            return lines.join("\\n");
        }
    """)
)

gb.configure_column(
    "SpeedUp",
    pinned="left",
    type=["textColumn"],
    headerTooltip=(
        "SpeedUp = True if all of:\n"
        "  • Running Days < 90\n"
        "  • AND Avg Intake Pressure > High intake threshold\n"
        "  • AND (Max Intake – Min Intake) < Small drawdown threshold\n"
        "  • AND Overload_Risk = False\n"
        "  • AND CapLoad > CapLoadPct threshold.\n"
        "Displays a green “✓” if True, blank if False."
    ),
    tooltipValueGetter=JsCode("""
        function(params) {
            // Only show tooltip if the cell value is '✗'
            if (params.value !== '✗') {
                return null;
            }
            var d = params.data;
            var lines = [];
            // Check each SpeedUp sub‐criterion via hidden Boolean fields:
            var speedC = [
                { name: "Running Days < 90",            ok: d.Speed_RunDays_OK     },
                { name: "Avg Intake > HighIntake",      ok: d.Speed_AvgIntake_OK   },
                { name: "Drawdown < SmallDrawdown",     ok: d.Speed_Drawdown_OK    },
                { name: "Overload_Risk is False",       ok: d.Speed_OverloadOK     },
                { name: "At_Max_Capacity is False",     ok: d.Speed_AtMaxOK        }
            ];
            speedC.forEach(function(c) {
                var status = c.ok ? "passed" : "failed";
                lines.push(c.name + ": " + status);
            });
            return lines.join("\\n");
        }
    """)
)

# Configure each derived column’s tooltip & formatting:
gb.configure_column(
    "At_Max_Capacity",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "At_Max_Capacity = Max Drive Amps ÷ Normal Running Amps.\n"
        "Cell is red if ≥ Cap load pct threshold, green otherwise."
    )
)
gb.configure_column(
    "Overload_Risk",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "Overload_Risk = Max Drive Amps ÷ Motor Overload.\n"
        "Cell is red if ≥ Risk pct threshold, green otherwise."
    )
)
gb.configure_column(
    "High Motor Temp",
    type=["numericColumn"],
    valueFormatter="x.toFixed(1)",
    headerTooltip=(
        "High Motor Temp = Max Motor Temp.\n"
        "Cell is red if ≥ High motor temp threshold, green otherwise."
    )
)
gb.configure_column(
    "High Downtime",
    type=["numericColumn"],
    valueFormatter="x.toFixed(1)",
    headerTooltip=(
        "High Downtime = average Downtime (Hr) over lookback window.\n"
        "Cell is red if > High downtime hrs threshold, green otherwise."
    )
)
gb.configure_column(
    "Max Vibration",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "Max Vibration = max(Avg Vib X, Avg Vib Y) over lookback.\n"
        "Cell is red if ≥ High vibration threshold, green otherwise."
    )
)
gb.configure_column(
    "Spread Ratio",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "SpreadRatio = (Max Drive Amps − Min Drive Amps) ÷ (Avg Drive Amps).\n"
        "If Min Drive Amps = 0 → displayed as N/A.\n"
        "Cell is red if ≥ Spread ratio threshold, green otherwise."
    )
)
gb.configure_column(
    "Tubing-Casing Δ",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "Tubing-Casing Δ = Avg Tubing Pressure − Avg Casing Pressure.\n"
        "Cell is red if ≤ Low Tub-Casing Δ threshold, green otherwise."
    )
)
gb.configure_column(
    "NearUnderload Ratio",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "NearUnderload Ratio = Avg Drive Amps ÷ Motor Underload.\n"
        "Cell is red if < Near-underload lower bound, green otherwise."
    )
)
gb.configure_column(
    "NearUnderload",
    type=["textColumn"],
    headerTooltip=(
        "True if (Avg Drive Amps ÷ Motor Underload) < Near-underload threshold.\n"
        "Displays a red “✗” if True, blank if False."
    )
)
gb.configure_column(
    "Normal_vs_Overload",
    type=["textColumn"],
    headerTooltip=(
        "True if (Latest Normal Running Amps) ≥ (Latest Motor Overload), else False.\n"
        "Displays a red “✗” if True, blank if False."
    )
)
gb.configure_column(
    "MissingSensor",
    type=["textColumn"],
    headerTooltip=(
        "True if Avg Motor Amps=0 or Avg Intake Pressure=0 or flat-line in either.\n"
        "Displays a red “✗” if True, blank if False."
    )
)
gb.configure_column(
    "Fault Count",
    type=["numericColumn"],
    valueFormatter="x.toFixed(0)",
    headerTooltip=(
        "Fault Count = cumulative faults over lookback window.\n"
        "Cell is red if ≥ High fault count threshold, green otherwise."
    )
)
gb.configure_column(
    "HighVib",
    type=["numericColumn"],
    valueFormatter="x.toFixed(2)",
    headerTooltip=(
        "HighVib = Max Vibration (max of Avg Vib X, Avg Vib Y).\n"
        "Cell is red if ≥ High vibration threshold, green otherwise."
    )
)
gb.configure_column(
    "Uptime %",
    type=["numericColumn"],
    valueFormatter="x == null ? '' : x.toFixed(1)",
    headerTooltip=(
        "Uptime % = mean uptime percentage over lookback window.\n"
        "Cell is red if < Low uptime threshold, green otherwise."
    )
)

gb.configure_column(
    "Drive Type",
    type=["textColumn"],
    headerTooltip="Drive Type (first value over lookback window)."
)
gb.configure_column(
    "State Detail/Op Mode",
    type=["textColumn"],
    headerTooltip="State Detail/Op Mode (first value over lookback window)."
)
# ─── Hide the helper Boolean fields from the visible grid ───
gb.configure_column("LowUptime_bool",       hide=True)
gb.configure_column("HighVib_bool",         hide=True)
gb.configure_column("SpreadFlag_bool",      hide=True)
gb.configure_column("HighMotorTemp_bool",   hide=True)
gb.configure_column("FaultHigh_bool",       hide=True)

gb.configure_column("Speed_RunDays_OK",     hide=True)
gb.configure_column("Speed_AvgIntake_OK",   hide=True)
gb.configure_column("Speed_Drawdown_OK",    hide=True)
gb.configure_column("Speed_OverloadOK",     hide=True)
gb.configure_column("Speed_AtMaxOK",        hide=True)


gb.configure_grid_options(enableBrowserTooltips=True)
gb.configure_default_column(resizable=True, minWidth=120)
gb.configure_grid_options(domLayout='normal')
gb.configure_columns(display_cols, cellStyle=js_color)
grid_opts = gb.build()

# ───────────── Page toggle ─────────────
page = st.sidebar.radio("View", ["Dashboard", "Raw Data"])

if page == "Dashboard":
    # 1) compute the four counts from df3:
    poor_count      = int(df3["PoorPerformance"].sum())
    speedup_count   = int(df3["SpeedUp"].sum())
    hightemp_count  = int(df3["HighMotorTemp"].sum())
    missing_count   = int(df3["MissingSensor"].sum())

    # 2) Create a row of 5 columns: one big for title/text, and four small for cards
    col_title, col_poor, col_speedup, col_ht, col_miss = st.columns([4, 1, 1, 1, 1])

    # 2a) Title + “Loaded X wells” goes in the first column
    with col_title:
        # Dashboard title
        st.markdown(
            "<h1 style='margin:0; "
            f"color: {'#ffffff' if night_mode else '#000000'};'>"
            "Daily Well-Performance Dashboard</h1>",
            unsafe_allow_html=True,
        )
        # “Loaded X wells from …” text
        date_range = (
            f"{min(last_dates)}"
            if hist_days == 1
            else f"{min(last_dates)} → {max(last_dates)}"
        )
        st.markdown(
            f"<p style='margin:0; font-size:14px; "
            f"color: {'#e0e0e0' if night_mode else '#333'};'>"
            f"Loaded <strong>{len(df3)}</strong> wells from {hist_days} day"
            f"{'s' if hist_days>1 else ''} ({date_range})</p>",
            unsafe_allow_html=True,
        )

    # 2b) Four cards, one per metric:

    # PoorPerformance (red)
    light_bg, dark_bg = "#FDE0E0", "#8B0000"
    light_txt, dark_txt = "#B00000", "#FFFFFF"
    with col_poor:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    Poor Performance
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {poor_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # SpeedUp (green)
    light_bg, dark_bg = "#E0FDE0", "#006400"
    light_txt, dark_txt = "#006400", "#FFFFFF"
    with col_speedup:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    Speed Up
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {speedup_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # HighMotorTemp (orange)
    light_bg, dark_bg = "#FFF1E0", "#CC8400"
    light_txt, dark_txt = "#CC6600", "#FFFFFF"
    with col_ht:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    High Motor Temp
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {hightemp_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # MissingSensor (purple)
    light_bg, dark_bg = "#F0E0FD", "#6A0080"
    light_txt, dark_txt = "#8000CC", "#FFFFFF"
    with col_miss:
        card_style = (
            f"background-color: {dark_bg}; color: {dark_txt};"
            if night_mode else
            f"background-color: {light_bg}; color: {light_txt};"
        )
        st.markdown(
            f"""
            <div style="
                {card_style}
                padding: 12px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0px 2px 4px rgba(0,0,0,0.15);
            ">
                <div style="font-size:14px; font-weight:600; margin-bottom:4px;">
                    Missing Sensor
                </div>
                <div style="font-size:24px; font-weight:bold;">
                    {missing_count}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # 3) Below the card row, show the filter & AG-Grid as before:
    st.markdown("")  # optional spacer
    flag = st.selectbox(
        "Filter wells by flag",
        [
            "All",
            "MissingSensor",
            "Normal_vs_Overload",
            "NearUnderload",
            "HighVib",
            "LowUptime",
            "PoorPerformance"
        ]
    )
    view = df_show if flag == "All" else df_show[df3[flag]]
    view = view.sort_values("TerribleScore", ascending=False)

    grid_theme = "ag-theme-alpine-dark" if night_mode else "ag-theme-alpine"
    AgGrid(
        view,
        gridOptions=grid_opts,
        allow_unsafe_jscode=True,
        use_container_width=True,
        fit_columns_on_grid_load=False,
        theme=grid_theme,
    )

else:
    # ─────────── “Raw Data” tab (unchanged) ───────────
    st.subheader("Full aggregated df3")
    if night_mode:
        st.dataframe(
            df3.style.set_properties(
                **{"background-color": "#2d2d2d", "color": "#e0e0e0", "border-color": "#555"}
            ),
            height=700,
        )
    else:
        st.dataframe(df3, height=700)

# ───────────── Export enriched CSV ─────────────
if st.button("Refresh & Save Enriched CSV"):
    outfile = pathlib.Path(f"{today}_well_report_enriched.csv")
    df_show.to_csv(outfile, index=False)
    st.success(f"Enriched CSV saved to **{outfile.resolve()}**")


# To run from PowerShell or Command Prompt:
# cd "C:\\Users\\Thai.phi\\OneDrive - Endurance Lift Solutions\\Desktop\\modbus dashboard"
# cd "C:\\Users\\Thai.phi\\OneDrive - Endurance Lift Solutions\\Desktop\\modbus dashboard"
# streamlit run "well review.py"
